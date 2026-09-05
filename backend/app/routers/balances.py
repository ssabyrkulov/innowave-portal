"""Снапшоты остатков: деньги по кассам/счетам и товары по складам.

Файлы 1С — это «остатки на сейчас», поэтому каждый импорт полностью
заменяет предыдущий снимок.
"""

import io
from collections import Counter
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import get_current_user

router = APIRouter(prefix="/balances", tags=["balances"])


def _num(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(" ", "").replace(",", ""))
    except ValueError:
        try:
            return float(str(value).replace(" ", "").replace(",", "."))
        except ValueError:
            return None


def _load_rows(content: bytes):
    """Строки первого листа книги.

    Под защитой весь разбор, а не только открытие: на некоторых файлах 1С
    ошибка возникала уже при чтении листа и уходила наружу «сырым» 500.
    Читаем без read_only — этот режим разбирает XML лениво и на файлах с
    нестандартной разметкой ведёт себя капризно."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))
    except HTTPException:
        raise
    except Exception as err:  # noqa: BLE001 — важен понятный ответ, не тип
        raise HTTPException(
            status_code=400,
            detail=f"Не удалось прочитать файл Excel: {type(err).__name__}: {err}",
        ) from err


def import_cash_balances_workbook(
    db: Session, content: bytes, filename: str, user_id: int,
    org: str = models.DEFAULT_ORG,
) -> dict:
    org = models.normalize_org(org)
    rows = _load_rows(content)
    # Два формата выгрузки. Старый — две колонки «Касса_Банк / СуммаОстаток».
    # Новый — «Счет | Место хранения | KGS | USD | RUB | Итого, сом»: остаток
    # разложен по валютам, а сомовый эквивалент лежит в «Итого, сом». Именно
    # он и нужен: касса в долларах на −6 500 USD — это −568 425 сом, и по
    # валютной колонке остаток был бы посчитан как шесть с половиной тысяч.
    header_idx, name_col, amount_col = None, 0, 1
    for i, row in enumerate(rows[:20]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        cset = {c for c in cells if c}
        if "Касса_Банк" in cset and "СуммаОстаток" in cset:
            header_idx = i
            name_col = cells.index("Касса_Банк")
            amount_col = cells.index("СуммаОстаток")
            break
        total = next((j for j, c in enumerate(cells)
                      if c.lower().startswith("итого")), None)
        if "Место хранения" in cset and total is not None:
            header_idx = i
            name_col = cells.index("Место хранения")
            amount_col = total
            break
    if header_idx is None:
        raise HTTPException(status_code=400, detail="Не найдены колонки остатков денег")

    # Как и с товарами: сначала разбираем, потом заменяем — иначе пустая
    # выгрузка стёрла бы остатки денег.
    parsed: list[tuple[str, float]] = []
    for row in rows[header_idx + 1:]:
        if not row or name_col >= len(row) or row[name_col] is None:
            continue
        account = str(row[name_col]).strip()
        # Новая выгрузка заканчивается строкой «ИТОГО». Сейчас она отсеивается
        # сама (место хранения там пусто), но если 1С однажды его заполнит,
        # итог приедет восьмым счётом и удвоит все деньги на портале.
        if account.lower().startswith(("итого", "всего")):
            continue
        amount = _num(row[amount_col] if amount_col < len(row) else None)
        if amount is None:
            continue
        parsed.append((account, amount))

    if not parsed:
        db.add(models.ImportLog(
            filename=f"[остатки денег, пусто] {filename}", user_id=user_id,
            added=0, skipped=0, errors_count=0,
        ))
        db.commit()
        return {
            "added": 0, "snapshot": True, "empty": True,
            "detail": "В файле нет строк остатков — прежние данные сохранены.",
        }

    db.query(models.CashBalance).filter(models.CashBalance.organization == org).delete()
    now = datetime.utcnow()
    for account, amount in parsed:
        db.add(models.CashBalance(
            account=account, amount=amount,
            organization=org, updated_at=now,
        ))

    db.add(models.ImportLog(
        filename=f"[остатки денег] {filename}", user_id=user_id,
        added=len(parsed), skipped=0, errors_count=0,
    ))
    db.commit()
    return {"added": len(parsed), "snapshot": True}


def _import_stock_matrix(db: Session, rows, header_idx: int,
                         filename: str, user_id: int, org: str) -> dict:
    """Матричный отчёт остатков: номенклатура × склады.

    Суммы в файле даны только итогом по номенклатуре — раскладываем их по
    складам пропорционально количеству, чтобы сумма складов сходилась с
    файлом копейка в копейку (через цену×количество сходилась бы только
    примерно: цена в файле округлена)."""
    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    itogo_idx = next((j for j, c in enumerate(header) if "ИТОГО" in c), None)
    price_idx = next((j for j, c in enumerate(header) if "Цена" in c), None)
    sum_idx = next((j for j, c in enumerate(header) if c == "Сумма"), None)
    if itogo_idx is None:
        raise HTTPException(status_code=400,
                            detail="В матричном отчёте не нашлась колонка ИТОГО")
    # НоменклатураGUID в матричном отчёте пока не приходит — остаток
    # единственная выгрузка без него, и позицию приходится узнавать по
    # названию. Колонку ждём заранее по двум причинам: с ней склейка
    # перестанет зависеть от написания, а без этой строчки она сломала бы
    # импорт — встав между номенклатурой и складами, GUID был бы прочитан
    # как ещё один склад с нечисловым остатком.
    guid_idx = next((j for j, c in enumerate(header)
                     if c and "GUID" in c.upper()), None)
    warehouses = [(j, header[j]) for j in range(1, itogo_idx)
                  if header[j] and j != guid_idx]

    parsed: list[models.StockBalance] = []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if not name or name == "ИТОГО":
            continue

        def cell(j):
            return row[j] if j is not None and j < len(row) else None

        total_qty = _num(cell(itogo_idx))
        total_sum = _num(cell(sum_idx))
        price = _num(cell(price_idx))
        per_wh = [(wh, _num(cell(j))) for j, wh in warehouses]
        per_wh = [(wh, q) for wh, q in per_wh if q]
        if not per_wh:
            continue  # позиция без остатков ни на одном складе
        for wh, q in per_wh:
            if total_sum is not None and total_qty:
                amount = total_sum * q / total_qty
            elif price is not None:
                amount = q * price
            else:
                amount = 0.0
            parsed.append(models.StockBalance(
                organization=org, product=name, warehouse=wh,
                product_guid=(str(cell(guid_idx) or "").strip().lower()
                              or None),
                qty=q, amount=round(amount, 2),
            ))

    if not parsed:
        db.add(models.ImportLog(
            filename=f"[остатки товаров, пусто] {filename}", user_id=user_id,
            added=0, skipped=0, errors_count=0,
        ))
        db.commit()
        return {"added": 0, "snapshot": True, "empty": True,
                "detail": "В файле нет строк остатков — прежние данные сохранены."}

    db.query(models.StockBalance).filter(
        models.StockBalance.organization == org).delete(synchronize_session=False)
    now = datetime.utcnow()
    for p in parsed:
        p.updated_at = now
        db.add(p)
    db.add(models.ImportLog(
        filename=f"[остатки товаров] {filename}", user_id=user_id,
        added=len(parsed), skipped=0, errors_count=0,
    ))
    db.commit()
    return {"added": len(parsed), "snapshot": True, "matrix": True}


def _import_stock_flat(db: Session, data_rows, header_cells: dict,
                       col_product: int, col_store: int,
                       filename: str, user_id: int, org: str) -> dict:
    """Плоская выгрузка остатков (наша обработка 1С): строка = товар на складе.

    Колонки ищутся по именам, лишние не мешают. НоменклатураGUID — по желанию:
    с ним товар связывается с SalesDoc по идентификатору, а не по названию."""
    col_qty = header_cells.get("КоличествоОстаток")
    col_amount = header_cells.get("СуммаОстаток")
    col_guid = header_cells.get("НоменклатураGUID")

    def cell(row, j):
        return row[j] if j is not None and j < len(row) else None

    parsed = []
    for row in data_rows:
        if not row:
            continue
        product = str(cell(row, col_product) or "").strip()
        if not product or product == "Итог":
            continue
        qty = _num(cell(row, col_qty))
        amount = _num(cell(row, col_amount))
        if qty is None and amount is None:
            continue
        parsed.append(models.StockBalance(
            organization=org,
            product=product,
            product_guid=str(cell(row, col_guid) or "").strip() or None,
            warehouse=str(cell(row, col_store) or "").strip() or None,
            qty=qty or 0,
            amount=amount or 0,
        ))

    if not parsed:
        db.add(models.ImportLog(
            filename=f"[остатки товаров, пусто] {filename}", user_id=user_id,
            added=0, skipped=0, errors_count=0,
        ))
        db.commit()
        return {"added": 0, "snapshot": True, "empty": True,
                "detail": "В файле нет строк остатков — прежние данные сохранены."}

    # Снимок: замена целиком в рамках организации (как у иерархического).
    db.query(models.StockBalance).filter(
        models.StockBalance.organization == org).delete(synchronize_session=False)
    now = datetime.utcnow()
    for p in parsed:
        p.updated_at = now
        db.add(p)
    db.add(models.ImportLog(
        filename=f"[остатки товаров] {filename}", user_id=user_id,
        added=len(parsed), skipped=0, errors_count=0,
    ))
    db.commit()
    return {"added": len(parsed), "snapshot": True, "flat": True}


def import_stock_balances_workbook(
    db: Session, content: bytes, filename: str, user_id: int,
    org: str = models.DEFAULT_ORG,
) -> dict:
    org = models.normalize_org(org)
    rows = _load_rows(content)
    header_idx = None
    for i, row in enumerate(rows[:20]):
        cells = {str(c).strip() for c in row if c is not None}
        if "СуммаОстаток" in cells and "КоличествоОстаток" in cells:
            header_idx = i
            break
    if header_idx is None:
        # Матричный отчёт по остаткам запасов (ВыгрузкаОстНью): строка —
        # номенклатура, колонки — склады, затем «ИТОГО кол», «Цена за ед»,
        # «Сумма». Третий формат остатков, который принимает портал.
        for i, row in enumerate(rows[:20]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if cells and cells[0] == "Номенклатура" \
                    and any("ИТОГО" in c for c in cells[1:]):
                return _import_stock_matrix(db, rows, i, filename, user_id, org)
        raise HTTPException(status_code=400, detail="Не найдены колонки остатков товаров")

    # Плоский формат нашей обработки: склад и номенклатура — отдельными
    # колонками, одна строка = товар на складе. Надёжнее иерархического
    # отчёта: не нужно угадывать, что в первой колонке — товар или склад.
    header_cells = {str(c).strip(): j for j, c in enumerate(rows[header_idx])
                    if c is not None}
    flat_product = next((header_cells[k] for k in
                         ("НоменклатураНаименование", "Номенклатура")
                         if k in header_cells), None)
    flat_store = next((header_cells[k] for k in
                       ("СкладНаименование", "Склад") if k in header_cells), None)
    if flat_product is not None and flat_store is not None:
        return _import_stock_flat(db, rows[header_idx + 1:], header_cells,
                                  flat_product, flat_store, filename, user_id, org)

    # Файл иерархический: строка товара, затем строки его складов.
    # Склад узнаём по справочнику складов из продаж этой организации.
    known_warehouses = {
        w for (w,) in db.query(models.Sale.warehouse)
        .filter(models.Sale.organization == org).distinct() if w
    }
    if not known_warehouses:  # продажи этой фирмы ещё не загружены — берём все
        known_warehouses = {w for (w,) in db.query(models.Sale.warehouse).distinct() if w}

    # Структурный запас: в иерархии склад повторяется под многими товарами, а
    # товар уникален. Имена, встречающиеся ≥2 раз, считаем складами — тогда
    # разбор не зависит от того, загружены ли уже продажи.
    data_rows = rows[header_idx + 1:]
    counts = Counter(
        str(r[0]).strip() for r in data_rows
        if r and r[0] is not None and str(r[0]).strip() not in ("", "Итог")
    )
    known_warehouses |= {n for n, c in counts.items() if c >= 2}

    # Сначала разбираем, и только потом трогаем базу: снимок заменяет данные
    # целиком, поэтому пустая выгрузка стёрла бы остатки подчистую.
    parsed: list[tuple[str, str, float, float]] = []
    current_product = None
    for row in data_rows:
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name == "Итог":
            continue
        amount = _num(row[1] if len(row) > 1 else None)
        qty = _num(row[2] if len(row) > 2 else None)
        if amount is None and qty is None:
            continue
        if name in known_warehouses and current_product:
            parsed.append((current_product, name, amount or 0, qty or 0))
        else:
            current_product = name

    if not parsed:
        # Файл без строк (в 1С выгрузили пустой отчёт) — прежние остатки
        # оставляем нетронутыми и честно говорим, что грузить было нечего.
        db.add(models.ImportLog(
            filename=f"[остатки товаров, пусто] {filename}", user_id=user_id,
            added=0, skipped=0, errors_count=0,
        ))
        db.commit()
        return {
            "added": 0, "snapshot": True, "empty": True,
            "detail": "В файле нет строк остатков — прежние данные сохранены. "
                      "Проверьте выгрузку в 1С (отчёт выгрузился пустым).",
        }

    db.query(models.StockBalance).filter(models.StockBalance.organization == org).delete()
    now = datetime.utcnow()
    for product, warehouse, amount, qty in parsed:
        db.add(models.StockBalance(
            product=product, warehouse=warehouse,
            amount=amount, qty=qty,
            organization=org, updated_at=now,
        ))
    added = len(parsed)

    db.add(models.ImportLog(
        filename=f"[остатки товаров] {filename}", user_id=user_id,
        added=added, skipped=0, errors_count=0,
    ))
    db.commit()
    return {"added": added, "snapshot": True}


@router.get("/cash")
def cash_balances(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    org: str = Query(default="all"),
):
    rows = models.org_scope(db.query(models.CashBalance), models.CashBalance, org).all()
    return {
        "total": round(sum(float(r.amount) for r in rows), 2),
        "updated_at": rows[0].updated_at.isoformat() if rows else None,
        "accounts": [
            {"account": r.account, "amount": float(r.amount)} for r in rows
        ],
    }


@router.get("/stock")
def stock_balances(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    org: str = Query(default="all"),
):
    rows = models.org_scope(db.query(models.StockBalance), models.StockBalance, org).all()
    return {
        "total_amount": round(sum(float(r.amount) for r in rows), 2),
        "total_qty": round(sum(float(r.qty) for r in rows), 3),
        "updated_at": rows[0].updated_at.isoformat() if rows else None,
        "items": [
            {
                "product": r.product,
                "warehouse": r.warehouse,
                "amount": float(r.amount),
                "qty": float(r.qty),
            }
            for r in rows
        ],
    }


@router.get("/stock-sources")
def stock_sources(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    org: str = Query(default="all"),
):
    """Остатки товаров из трёх источников рядом: управленка, налоговая, SalesDoc.

    Источники разной природы, и это принципиально для чтения таблицы:

    * **управленка** — снапшот «Остатки товаров» из 1С, факт на дату выгрузки;
    * **налоговая** — снапшота нет, в пакете налоговой такого файла вообще не
      выгружается. Считаем из движений: поступления, оприходования и
      возвраты от покупателей в плюс, реализации, списания и возвраты
      поставщику в минус. Это «как должно быть по документам», а не факт
      склада;
    * **SalesDoc** — остатки торговых точек из зеркала, только количество:
      сумм SalesDoc не отдаёт.

    Сходиться они не обязаны и в норме не сойдутся: налоговый контур ведёт не
    весь товар, а в SalesDoc лежат остатки точек, а не своих складов. Смысл
    таблицы — видеть все три числа рядом и замечать, когда расхождение
    выходит за привычное.
    """
    from .purchases import _group_of, _norm_product, _size_of
    from .salesdoc import _store_ids_for_org

    # Ключ сопоставления — нормализованное имя номенклатуры. Названия в трёх
    # системах пишутся по-разному, но после нормализации совпадают.
    rows: dict[str, dict] = {}

    def cell(name: str | None) -> dict:
        key = _norm_product(name)
        e = rows.get(key)
        if e is None:
            e = rows[key] = {"product": name or "—", "upr": None,
                             "nal": None, "sd": None, "upr_amount": None}
        return e

    # --- Управленка: снапшот 1С ---
    upr_rows = models.org_scope(
        db.query(models.StockBalance), models.StockBalance, org).all()
    guid_key: dict[str, str] = {}
    for r in upr_rows:
        e = cell(r.product)
        e["upr"] = (e["upr"] or 0.0) + float(r.qty or 0)
        e["upr_amount"] = (e["upr_amount"] or 0.0) + float(r.amount or 0)
        if r.product_guid:
            guid_key[str(r.product_guid)] = _norm_product(r.product)

    # --- Налоговая: расчёт из движений ---
    tax_q = models.org_scope(
        db.query(models.TaxOperation), models.TaxOperation, org)
    # Приход: поступление, оприходование (излишки, ввод остатков), возврат от
    # покупателя. Расход: реализация, списание, возврат поставщику.
    # Оприходований и возвратов поставщику здесь раньше не было, и колонка
    # врала в обе стороны: излишки не прибавлялись, возвраты поставщику не
    # вычитались. Состав движений теперь тот же, что в сверке контуров на
    # главной, — иначе два блока на одной странице показывали разное.
    # Перемещения не в счёт: они двигают товар между своими складами.
    SIGN = {"purchase": 1, "stock_in": 1, "return": 1,
            "sale": -1, "writeoff": -1, "return_supplier": -1}
    for op in tax_q.filter(models.TaxOperation.kind.in_(tuple(SIGN))).all():
        if not op.product:
            continue
        # Склад двигают только товарные документы. «Поступление услуги» —
        # это услуги, а «Поступление доп расходов» перечисляет ТЕ ЖЕ товары
        # ещё раз, чтобы разнести на них таможню и доставку. Без этого
        # отсева остаток Хайджина завышался на 856 583 шт из 635 066
        # закупленных — больше чем вдвое.
        src = (op.source or "").lower()
        if "услуг" in src or "доп расход" in src:
            continue
        # Товары для перепродажи — счета 161x–164x. Бензин, дизтопливо и
        # мебель приходуются тем же документом «Поступление товары», но на
        # счета материалов (1710) и МБП (1750): в остатках товаров им не
        # место, а по количеству от подгузников их не отличить.
        acc = str(op.account or "")
        if acc and not acc.startswith(("161", "162", "163", "164")):
            continue
        e = cell(op.product)
        e["nal"] = (e["nal"] or 0.0) + SIGN[op.kind] * float(op.qty or 0)

    # --- SalesDoc: зеркало остатков точек ---
    sd_q = db.query(models.SalesDocStock)
    store_ids = _store_ids_for_org(db, org)
    if store_ids:
        sd_q = sd_q.filter(models.SalesDocStock.store_sd_id.in_(store_ids))
    # Склады без привязки к фирме попадают в обе — так задумано в сверке,
    # чтобы отгрузки не исчезали молча. Но в остатках это значит, что одни и
    # те же штуки видны и у Хайджина, и у Инновейва. Считаем их отдельно,
    # чтобы в карточке было честно написано, сколько числа спорны.
    mapped = {str(st.store_id).lower() for st in db.query(models.SalesDocStore).all()
              if st.store_id and st.organization}
    sd_unmapped_qty = 0.0
    sd_unmapped_stores: set = set()
    for r in sd_q.all():
        # Код 1С точнее названия: позицию могли переименовать в одной системе
        # и не переименовать в другой.
        key = guid_key.get(str(r.code_1c or "")) if r.code_1c else None
        e = rows.get(key) if key else None
        if e is None:
            e = cell(r.product_name)
        qty = float(r.quantity or 0)
        e["sd"] = (e["sd"] or 0.0) + qty
        if str(r.store_sd_id or "").lower() not in mapped:
            sd_unmapped_qty += qty
            sd_unmapped_stores.add(r.store_sd_id)

    items = []
    for e in rows.values():
        if not any(abs(e[k] or 0) >= 0.001 for k in ("upr", "nal", "sd")):
            continue
        items.append({
            "product": e["product"],
            "upr": None if e["upr"] is None else round(e["upr"], 1),
            "nal": None if e["nal"] is None else round(e["nal"], 1),
            "sd": None if e["sd"] is None else round(e["sd"], 1),
            "upr_amount": None if e["upr_amount"] is None else round(e["upr_amount"], 2),
        })
    items.sort(key=lambda x: (_group_of(x["product"]), _size_of(x["product"]),
                              x["product"]))

    def total(field: str) -> float:
        return round(sum(i[field] or 0 for i in items), 1)

    synced = None
    try:
        from ..services import salesdoc_mirror
        synced = salesdoc_mirror.status(db).get("synced_at")
    except Exception:  # noqa: BLE001 — зеркало не должно ронять карточку
        synced = None

    return {
        "org": org,
        "sources": {
            "upr": {
                "label": "Управленка",
                "note": "снапшот 1С",
                "available": bool(upr_rows),
                "total_qty": total("upr"),
                "total_amount": round(sum(i["upr_amount"] or 0 for i in items), 2),
                "updated_at": (upr_rows[0].updated_at.isoformat()
                               if upr_rows else None),
            },
            "nal": {
                "label": "Налоговая",
                "note": "расчёт из движений",
                "available": any(i["nal"] is not None for i in items),
                "total_qty": total("nal"),
            },
            "sd": {
                "label": "SalesDoc",
                "note": "остатки точек",
                "available": any(i["sd"] is not None for i in items),
                "total_qty": total("sd"),
                "synced_at": synced,
                "unmapped_stores": len(sd_unmapped_stores),
                "unmapped_qty": round(sd_unmapped_qty, 1),
            },
        },
        "items": items,
    }
