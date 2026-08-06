import hashlib
import io
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from fastapi import (
    APIRouter,
    Depends,
    Form,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import get_current_user, require_roles

router = APIRouter(prefix="/sales", tags=["sales"])

can_import = require_roles(models.Role.admin, models.Role.accountant)
admin_only = require_roles(models.Role.admin)

# Заголовки колонок 1С → поля модели. Ищем по именам, а не по позициям,
# чтобы перестановка колонок в выгрузке не ломала импорт.
HEADER_MAP = {
    "Дата": "date",
    "КонтрагентНаименование": "client",
    "Склад": "warehouse",
    "НоменклатураНаименование": "product",
    "Количество": "qty",
    "Цена": "price",
    "Сумма": "amount",
    "ВалютаДокументаНаименование": "currency",
    "Номер": "doc_number",
    "СуммаДокумента": "doc_total",
    "НоменклатураЕдиницаИзмерения": "unit",
    "КонтрагентSD_АгентНаименование": "agent",
    "ПроцентСкидкиНаценки": "discount_pct",
    "СчетУчета": "account",
    "ОтветственныйНаименование": "responsible",
    # Новый вариант выгрузки: ФИО того, кто оформил документ
    "ОтветственныйФИО": "responsible",
    # GUID документа: появился в обновлённых выгрузках 1С. Колонка
    # необязательная — файлы без неё грузятся как раньше.
    "ДокументGUID": "doc_guid",
}

REQUIRED_FIELDS = {"date", "client", "product", "qty", "price", "amount"}


def _parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.split()[0], "%d.%m.%Y").date()
        except ValueError:
            return None
    return None


def _parse_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def _row_hash(d: dict, org: str = models.DEFAULT_ORG) -> str:
    key = "|".join(
        str(d.get(f, ""))
        for f in (
            "date", "client", "product", "qty", "price",
            "amount", "doc_number", "warehouse", "doc_total",
        )
    )
    return hashlib.sha256(f"{org}|{key}".encode()).hexdigest()


@router.post("/import")
async def import_sales(
    file: UploadFile,
    replace_period: bool = Form(default=False),
    org: str = Form(default=models.DEFAULT_ORG),
    db: Session = Depends(get_db),
    current: models.User = Depends(can_import),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ожидается файл Excel (.xlsx)",
        )
    content = await file.read()
    return import_sales_workbook(
        db, content, file.filename, current.id, replace_period, org
    )


def parse_sales_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    """Разбирает Excel продаж (или возвратов в том же формате) в строки.

    Возвращает (parsed_rows, errors). Используется импортом продаж и
    очисткой ошибочно загруженных возвратов.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Не удалось открыть файл — убедитесь, что это Excel (.xlsx)",
        )

    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    # Ищем строку заголовков в первых 20 строках
    header_idx = None
    columns: dict[int, str] = {}
    buffered = []
    for i, row in enumerate(rows):
        buffered.append(row)
        if i >= 20:
            break
        matched = {
            j: HEADER_MAP[str(cell).strip()]
            for j, cell in enumerate(row)
            if cell is not None and str(cell).strip() in HEADER_MAP
        }
        if "date" in matched.values() and "amount" in matched.values():
            header_idx = i
            columns = matched
            break

    if header_idx is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Не найдена строка заголовков (ожидаются колонки 1С: "
            "Дата, КонтрагентНаименование, Сумма и т.д.)",
        )

    # --- Фаза 1: разбор всех строк файла ---
    parsed_rows: list[dict] = []
    errors: list[str] = []

    def process(row, line_no):
        data = {}
        for j, field in columns.items():
            data[field] = row[j] if j < len(row) else None
        if all(v is None for v in data.values()):
            return
        parsed = {
            "date": _parse_date(data.get("date")),
            "client": str(data.get("client") or "").strip() or None,
            "warehouse": str(data.get("warehouse") or "").strip() or None,
            "product": str(data.get("product") or "").strip() or None,
            "qty": _parse_float(data.get("qty")),
            "price": _parse_float(data.get("price")),
            "amount": _parse_float(data.get("amount")),
            "currency": str(data.get("currency") or "KGS").strip()[:3] or "KGS",
            "doc_number": str(data.get("doc_number") or "").strip() or None,
            "doc_total": _parse_float(data.get("doc_total")),
            "unit": str(data.get("unit") or "").strip() or None,
            "agent": str(data.get("agent") or "").strip() or None,
            "discount_pct": _parse_float(data.get("discount_pct")),
            "account": str(data.get("account") or "").strip() or None,
            "responsible": str(data.get("responsible") or "").strip() or None,
            "doc_guid": str(data.get("doc_guid") or "").strip() or None,
        }
        missing = [f for f in REQUIRED_FIELDS if parsed.get(f) is None]
        if missing:
            if len(errors) < 20:
                errors.append(f"Строка {line_no}: нет значений {', '.join(missing)}")
            return
        parsed_rows.append(parsed)

    line_no = header_idx + 1
    for line_no, row in enumerate(buffered[header_idx + 1:], start=header_idx + 2):
        process(row, line_no)
    for line_no, row in enumerate(rows, start=line_no + 1):
        process(row, line_no)

    return parsed_rows, errors


def import_sales_workbook(
    db: Session,
    content: bytes,
    filename: str,
    user_id: int,
    replace_period: bool = False,
    org: str = models.DEFAULT_ORG,
) -> dict:
    """Импорт выгрузки продаж из байтов Excel (используется и веб-загрузкой,
    и автоприёмом из Google Drive)."""
    org = models.normalize_org(org)
    parsed_rows, errors = parse_sales_workbook(content)

    replaced = 0
    if replace_period and parsed_rows:
        # Заменяем период ТОЛЬКО в рамках своей организации: правки задним
        # числом в 1С корректно попадут в базу, а не останутся дублями.
        dmin = min(p["date"] for p in parsed_rows)
        dmax = max(p["date"] for p in parsed_rows)
        replaced = (
            db.query(models.Sale)
            .filter(models.Sale.organization == org,
                    models.Sale.date >= dmin, models.Sale.date <= dmax)
            .delete(synchronize_session=False)
        )
        db.flush()

    # --- Фаза 2: вставка с защитой от дублей (хэш учитывает организацию) ---
    existing_hashes = {h for (h,) in db.query(models.Sale.row_hash).all()}
    seen_in_file: set[str] = set()
    added = 0
    skipped = 0
    for parsed in parsed_rows:
        h = _row_hash(parsed, org)
        if h in existing_hashes or h in seen_in_file:
            skipped += 1
            continue
        seen_in_file.add(h)
        db.add(models.Sale(**parsed, organization=org, row_hash=h))
        added += 1

    db.add(models.ImportLog(
        filename=filename or "upload.xlsx",
        user_id=user_id,
        added=added,
        skipped=skipped,
        errors_count=len(errors),
        replace_period=replace_period,
    ))
    db.commit()
    total = db.query(models.Sale).count()
    return {
        "added": added,
        "skipped_duplicates": skipped,
        "replaced_rows": replaced,
        "errors": errors,
        "total_in_db": total,
    }


# Документная реализация (Innowave): Дата / Сумма / Валюта / Контрагент.
DOC_HEADERS = {"Дата": "date", "Сумма": "amount", "Валюта": "currency",
               "Контрагент": "client", "ДокументGUID": "doc_guid"}


def import_sales_docs_workbook(
    db: Session, content: bytes, filename: str, user_id: int,
    org: str = models.DEFAULT_ORG,
) -> dict:
    """Импорт документной реализации (одна строка = документ с итогом, без
    номенклатуры). Товар подставляем плейсдержателем, кол-во=1, цена=сумма —
    чтобы дебиторка и обороты считались, а товарные проверки такие строки
    игнорировали."""
    org = models.normalize_org(org)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось открыть файл Excel")

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header_idx, columns, buffered = None, {}, []
    for i, row in enumerate(rows_iter):
        buffered.append(row)
        if i >= 20:
            break
        matched = {
            j: DOC_HEADERS[str(c).strip()]
            for j, c in enumerate(row)
            if c is not None and str(c).strip() in DOC_HEADERS
        }
        if {"date", "amount", "client"} <= set(matched.values()):
            header_idx, columns = i, matched
            break
    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Не найдены колонки документной реализации (Дата, Сумма, "
                   "Валюта, Контрагент)",
        )

    parsed, errors = [], []

    def process(row, line_no):
        data = {f: (row[j] if j < len(row) else None) for j, f in columns.items()}
        if all(v is None for v in data.values()):
            return
        dt = _parse_date(data.get("date"))
        amount = _parse_float(data.get("amount"))
        client = str(data.get("client") or "").strip()
        if not dt or amount is None or not client:
            if len(errors) < 20:
                errors.append(f"Строка {line_no}: нет даты, суммы или контрагента")
            return
        parsed.append({
            "src_dt": str(data.get("date")),
            "date": dt,
            "amount": amount,
            "currency": (str(data.get("currency") or "KGS").strip() or "KGS")[:3],
            "client": client,
        })

    line_no = header_idx + 1
    for line_no, row in enumerate(buffered[header_idx + 1:], start=header_idx + 2):
        process(row, line_no)
    for line_no, row in enumerate(rows_iter, start=line_no + 1):
        process(row, line_no)

    existing = {h for (h,) in db.query(models.Sale.row_hash).all()}
    seen: set[str] = set()
    occurrences: dict[str, int] = defaultdict(int)
    added = skipped = 0
    for p in parsed:
        base = "|".join(str(p[f]) for f in ("src_dt", "amount", "currency", "client"))
        occurrences[base] += 1
        h = hashlib.sha256(f"{org}|{base}#{occurrences[base]}".encode()).hexdigest()
        if h in existing or h in seen:
            skipped += 1
            continue
        seen.add(h)
        db.add(models.Sale(
            date=p["date"], client=p["client"], product=models.DOC_SALE_PRODUCT,
            qty=1, price=p["amount"], amount=p["amount"], currency=p["currency"],
            doc_total=p["amount"], organization=org, row_hash=h,
        ))
        added += 1

    db.add(models.ImportLog(
        filename=f"[реализация-док] {filename}", user_id=user_id,
        added=added, skipped=skipped, errors_count=len(errors),
    ))
    db.commit()
    return {
        "added": added,
        "skipped_duplicates": skipped,
        "errors": errors,
        "total_in_db": db.query(models.Sale).count(),
    }


@router.get("/imports")
def list_imports(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    limit: int = Query(default=20, le=100),
):
    logs = (
        db.query(models.ImportLog)
        .order_by(models.ImportLog.created_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id": l.id,
            "filename": l.filename,
            "user": l.user.full_name if l.user else None,
            "added": l.added,
            "skipped": l.skipped,
            "errors_count": l.errors_count,
            "replace_period": l.replace_period,
            "created_at": l.created_at.isoformat(),
        }
        for l in logs
    ]


@router.get("/summary")
def sales_summary(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    top: int = Query(default=10, le=50),
    org: str = Query(default="all"),
):
    query = models.org_scope(db.query(models.Sale), models.Sale, org)
    if date_from:
        query = query.filter(models.Sale.date >= date_from)
    if date_to:
        query = query.filter(models.Sale.date <= date_to)
    sales = query.all()

    revenue = 0.0
    docs: set[str] = set()
    clients: dict[str, dict] = {}
    products: dict[str, dict] = {}
    agents: dict[str, dict] = {}
    monthly: dict[str, dict] = {}
    min_date = max_date = None

    for s in sales:
        amt = float(s.amount)
        revenue += amt
        if s.doc_number:
            docs.add(f"{s.doc_number}|{s.date}")
        c = clients.setdefault(s.client, {"name": s.client, "revenue": 0.0, "docs": set()})
        c["revenue"] += amt
        if s.doc_number:
            c["docs"].add(s.doc_number)
        p = products.setdefault(s.product, {"name": s.product, "revenue": 0.0, "qty": 0.0})
        p["revenue"] += amt
        p["qty"] += float(s.qty)
        # В старой выгрузке продавца несёт колонка агента, в новой — ФИО
        # ответственного; берём что есть.
        agent_name = s.agent or s.responsible
        if agent_name:
            a = agents.setdefault(agent_name, {"name": agent_name, "revenue": 0.0, "docs": set()})
            a["revenue"] += amt
            if s.doc_number:
                a["docs"].add(s.doc_number)
        mkey = s.date.strftime("%Y-%m")
        m = monthly.setdefault(mkey, {"month": mkey, "revenue": 0.0, "docs": set()})
        m["revenue"] += amt
        if s.doc_number:
            m["docs"].add(s.doc_number)
        if min_date is None or s.date < min_date:
            min_date = s.date
        if max_date is None or s.date > max_date:
            max_date = s.date

    def finalize(items, sort_key="revenue"):
        out = []
        for it in items:
            it = dict(it)
            if isinstance(it.get("docs"), set):
                it["docs"] = len(it["docs"])
            out.append(it)
        return sorted(out, key=lambda x: -x[sort_key])

    return {
        "lines": len(sales),
        "revenue": revenue,
        "docs": len(docs),
        "clients": len(clients),
        "avg_doc": revenue / len(docs) if docs else 0,
        "date_from": min_date.isoformat() if min_date else None,
        "date_to": max_date.isoformat() if max_date else None,
        "monthly": sorted(
            ({"month": m["month"], "revenue": m["revenue"], "docs": len(m["docs"])}
             for m in monthly.values()),
            key=lambda x: x["month"],
        ),
        "top_clients": finalize(clients.values())[:top],
        "top_products": finalize(products.values())[:top],
        "top_agents": finalize(agents.values())[:top],
    }


@router.get("/products")
def sales_products(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    q: str | None = Query(default=None, description="Поиск по названию номенклатуры"),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    limit: int = Query(default=200, le=1000),
    org: str = Query(default="all"),
):
    """Продажи в разрезе номенклатуры: сколько штук и на какую сумму продано
    каждого товара за период. Поддерживает поиск по подстроке названия —
    под вопросы вида «сколько продали туалетной бумаги по сегодня».
    """
    query = models.org_scope(db.query(models.Sale), models.Sale, org)
    if date_from:
        query = query.filter(models.Sale.date >= date_from)
    if date_to:
        query = query.filter(models.Sale.date <= date_to)
    if q:
        # регистронезависимый поиск по подстроке (работает и в SQLite, и в PG)
        query = query.filter(models.Sale.product.ilike(f"%{q.strip()}%"))
    sales = query.all()

    products: dict[str, dict] = {}
    total_qty = total_amount = 0.0
    total_docs: set[str] = set()
    for s in sales:
        amt = float(s.amount)
        qty = float(s.qty)
        total_qty += qty
        total_amount += amt
        if s.doc_number:
            total_docs.add(f"{s.doc_number}|{s.date}")
        p = products.setdefault(s.product, {
            "product": s.product, "qty": 0.0, "amount": 0.0,
            "unit": s.unit or "шт", "docs": set(),
            "first_date": s.date, "last_date": s.date,
        })
        p["qty"] += qty
        p["amount"] += amt
        if s.unit and p["unit"] == "шт":
            p["unit"] = s.unit
        if s.doc_number:
            p["docs"].add(f"{s.doc_number}|{s.date}")
        if s.date < p["first_date"]:
            p["first_date"] = s.date
        if s.date > p["last_date"]:
            p["last_date"] = s.date

    rows = sorted(
        (
            {
                "product": p["product"],
                "qty": round(p["qty"], 3),
                "amount": round(p["amount"], 2),
                "unit": p["unit"],
                "docs": len(p["docs"]),
                "first_date": p["first_date"].isoformat(),
                "last_date": p["last_date"].isoformat(),
            }
            for p in products.values()
        ),
        key=lambda x: -x["amount"],
    )

    return {
        "query": q or "",
        "count": len(rows),
        "total_qty": round(total_qty, 3),
        "total_amount": round(total_amount, 2),
        "total_docs": len(total_docs),
        "products": rows[:limit],
    }


@router.delete("", status_code=status.HTTP_204_NO_CONTENT)
def clear_sales(
    db: Session = Depends(get_db),
    _: models.User = Depends(admin_only),
):
    """Полная очистка данных продаж (только админ) — для переимпорта."""
    db.query(models.Sale).delete()
    db.commit()
