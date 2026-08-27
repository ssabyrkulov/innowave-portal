"""Автоприём выгрузок 1С, которые Google Apps Script пушит из папки Drive.

Скрипт (docs/AUTOSYNC.md) раз в N минут отправляет новые/изменённые файлы
на POST /integrations/inbox с токеном. Тип файла определяется по заголовкам
колонок, дубли отсекает построчная дедупликация импортёров.
"""

import hashlib
import io
import secrets
import traceback

import openpyxl
from fastapi import APIRouter, Depends, Form, Header, HTTPException, UploadFile
from sqlalchemy.orm import Session

from .. import models
from ..services import xlsx
from ..config import settings
from ..database import get_db
from ..deps import get_current_user, require_roles
from ..security import hash_password
from .balances import (
    import_cash_balances_workbook,
    import_stock_balances_workbook,
)
from .expenses import import_expenses_workbook
from .receipts import import_receipts_workbook
from .returns import import_return_lines_workbook, import_returns_workbook
from .sales import (
    import_sales_docs_workbook,
    import_sales_workbook,
)

router = APIRouter(prefix="/integrations", tags=["integrations"])

ROBOT_EMAIL = "robot@innowave.portal"


def _recover_filename(name: str) -> str:
    """Чинит имя файла, если оно пришло «моджибейком» (UTF-8 байты,
    прочитанные как latin-1) — типичная беда multipart-загрузок с
    кириллицей. Для корректных кириллических имён encode('latin-1')
    падает → возвращаем как есть.
    """
    try:
        return name.encode("latin-1").decode("utf-8")
    except (UnicodeError, AttributeError):
        return name


def _require_token(authorization: str | None) -> None:
    if not settings.inbox_token:
        raise HTTPException(
            status_code=503,
            detail="Автоприём выключен: не задан INBOX_TOKEN на сервере",
        )
    # Сравниваем байты, а не строки: secrets.compare_digest на str требует
    # чистого ASCII и на любом другом символе бросает TypeError — запрос падал
    # с «500 Внутренняя ошибка» вместо честного «неверный токен», из-за чего
    # причина выглядела как сбой разбора файла.
    expected = f"Bearer {settings.inbox_token}"
    ok = bool(authorization) and secrets.compare_digest(
        (authorization or "").encode("utf-8"), expected.encode("utf-8")
    )
    if not ok:
        raise HTTPException(
            status_code=401,
            detail="Неверный токен автоприёма: проверьте TOKEN в скрипте "
                   "Google Apps Script и INBOX_TOKEN на сервере (частая "
                   "причина — лишний пробел или кириллическая буква при "
                   "копировании).",
        )


def _robot_user(db: Session) -> models.User:
    """Служебный пользователь для журнала автозагрузок (вход запрещён)."""
    user = db.query(models.User).filter_by(email=ROBOT_EMAIL).first()
    if user is None:
        user = models.User(
            email=ROBOT_EMAIL,
            full_name="Автозагрузка (Drive)",
            role=models.Role.viewer,
            is_active=False,
            hashed_password=hash_password(secrets.token_urlsafe(24)),
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    return user


# Каноничный файл на каждый тип данных. У некоторых выгрузок есть несколько
# вариантов (Реал/Реал2, Возв/ТовВозв) — это разные форматы одних и тех же
# данных, которые тестировали с 1С. Грузим ТОЛЬКО каноничный, дубли-варианты
# явно пропускаем, чтобы не задваивать выручку/возвраты.
# Чтобы сменить каноничный файл — поправьте условия ниже.
def org_from_name(filename: str) -> str | None:
    """Фирма по имени файла: «Хайджин_…» / «Инновейв_…». None — не понять.

    Раньше фирму определяла только папка Drive. Это работало, пока папок было
    две, но выгрузки четырёх баз лежат в одной, а фирма написана в имени файла
    первым словом. Имя надёжнее папки ещё и потому, что файл, случайно
    положенный не туда, уедет в чужую фирму молча, а по имени приедет куда
    назван.

    Важно: «Хайджин» проверяем ПЕРВЫМ. Полное название фирмы — «Инновейв
    Хайджин», и по подстроке «инновейв» она бы досталась второй фирме.
    """
    name = (filename or "").lower()
    if "хайджин" in name or "hygiene" in name or "haydzhin" in name:
        return "hygiene"
    if "инновейв" in name or "innowave" in name or "innovejv" in name:
        return "innowave"
    return None


# Виды выгрузок, которые 1С кладёт в папку, а портал пока не ведёт. Список
# нужен не для маршрутизации (её делает classify_by_name), а для отчёта: те
# виды, что двигают склад, — прямая причина расхождения расчётного остатка с
# 1С. Документ товар подвинул, а наша математика его не видела, потому что
# файл был пропущен молча. Флаг moves_stock отделяет их от справочников и
# бухгалтерских регистров, которые на остаток не влияют вообще.
UNSUPPORTED_KINDS: tuple[tuple[str, tuple[str, ...], bool, str], ...] = (
    # «Возврат товаров поставщику» выгружается только по налоговому контуру;
    # управленческого файла нет, поэтому расход со склада по нему не виден.
    ("Возврат товаров поставщику",
     ("поставщик", "postavshik", "postavshchik"), True,
     "расход со склада, импортёра нет — расчётный остаток завышен"),
    # Инвентаризация в 1С проводок не делает: она фиксирует отклонение, а
    # товар двигают созданные на её основании оприходование (излишек) и
    # списание (недостача). Оба загружаются, поэтому считать ещё и её
    # значило бы задвоить движение.
    ("Инвентаризация", ("инвентариз", "inventariz"), False,
     "проводок не делает: излишки идут оприходованием, недостачи списанием"),
    ("Перемещение товаров", ("перемещен", "peremeshen", "peremeshch"), False,
     "между складами одной фирмы; расчёт ведётся по фирме целиком"),
    ("Движение МБП", ("движение мбп", "dvijenie mbp", "dvizhenie mbp"), False,
     "малоценка учитывается отдельно от товаров"),
    ("Корректировка долга", ("корректировка долга", "korrektirovka dolga",
                             "korrektirovka doljna"), False,
     "денежная корректировка взаиморасчётов, товара не касается"),
    ("Корректировка", ("корректировк", "korrektirovk"), True,
     "меняет уже проведённый документ — может задевать товар"),
    ("Взаимозачёт", ("взаимозач", "vzaimozach"), False, ""),
    ("Авансовый отчёт", ("авансов", "avansov", "подотчет", "подотчёт",
                         "podotchet"), False, ""),
    ("Счёт на оплату", ("счет на оплату", "счёт на оплату",
                        "schet na oplatu"), False, ""),
    ("Оборотно-сальдовая ведомость", ("оборотно", "oborotno"), False, ""),
    ("ГТД", ("гтд", "gtd"), False, ""),
    ("Журнал проводок", ("журнал проводок", "jurnal provodok",
                         "zhurnal provodok"), False, ""),
    ("Конвертация", ("конвертац", "konvertac"), False, ""),
    ("Начисление зарплаты", ("начисление зарплат", "nachislenie zarplat"),
     False, ""),
    ("Проблемные документы", ("проблемные документ", "problemnye dokument"),
     False, ""),
    ("Ручные операции", ("ручные операц", "ruchnye operac"), False, ""),
    ("ЭСФ / счета-фактуры", ("эсф", "esf", "счет-фактур", "счёт-фактур",
                             "schet-faktur", "бланки счетов",
                             "blanki schetov"), False, ""),
    ("Дополнительные расходы", ("дополнительн", "dopolnitel"), False, ""),
)


def unsupported_kind(filename: str) -> tuple[str, bool, str]:
    """Имя вида, признак «двигает склад» и пояснение — по имени файла.

    Порядок в таблице значим: «Корректировка долга» обязана стоять раньше
    общей «Корректировки», иначе денежный документ попадёт в товарные и
    будет зря пугать в отчёте."""
    name = (filename or "").lower()
    for label, tokens, moves, note in UNSUPPORTED_KINDS:
        if any(t in name for t in tokens):
            return label, moves, note
    return "Прочее", False, ""


def classify_by_name(filename: str, org: str = models.DEFAULT_ORG) -> str | None:
    """Тип выгрузки по имени файла — самый надёжный сигнал. None → sniff_kind.

    Каноничные файлы зависят от организации: у Hygiene реализация = Реал2, а
    возвраты — построчный ТовВозв; у Innowave выгружается только Реал (без «2»)
    и документный Возв. Поэтому логика ветвится по org.
    """
    name = (filename or "").lower()
    if name.startswith("~$"):
        return "ignore"  # временный файл Excel

    def has(*tokens: str) -> bool:
        return any(t in name for t in tokens)

    # Налоговый контур. Файл с меткой налоговой уходит своим импортёром в
    # отдельную таблицу и с управленческими не смешивается: имя
    # «Хайджин_НАЛОГОВАЯ_Реализация товары» иначе распозналось бы как обычная
    # реализация и задвоило бы продажи вдвое. Короткую метку ловим только в
    # скобках и подчёркиваниях: голое «nal» встречается внутри обычных слов
    # (analiz, nalichnie), и как подстрока опасно.
    if has("налог", "nalog", "[nal]", "[нал]", "_nal_", "_нал_"):
        return "tax"

    # Корректировки — раньше всех предметных правил. «Корректировка
    # реализации» содержит «реализац» и уезжала в импортёр продаж, а
    # «Корректировка поступления товаров» — в импортёр закупок. Это не просто
    # чужой файл в чужой таблице: корректировка меняет уже проведённый
    # документ, и загруженная как самостоятельная продажа она завышает
    # проданное, а как закупка — поступившее. Оба перекоса бьют ровно по той
    # арифметике, которой считается расчётный остаток.
    if has("корректировк", "korrektirovk"):
        return "unsupported"

    # --- Новая схема имён: «Фирма_Управленка_ТипВыгрузки» полными словами ---
    # Понимаем кириллицу и транслит. Эти правила стоят РАНЬШЕ старых коротких
    # токенов, и это важно: «Реализация» иначе попала бы в правило «реал»,
    # которое для Hygiene считает файл без «2» дублем и молча пропускает.
    if has("реализац", "realizac"):
        # Формат (построчный/документный) импортёр продаж определяет сам
        # по содержимому — одно имя работает для обеих фирм.
        return "sales"
    if has("возврат", "vozvrat"):
        # «Возврат товаров ПОСТАВЩИКУ» — это закупки, портал их не ведёт;
        # клиентский возврат в 1С называется «…от покупателя».
        if has("поставщик", "postavshik", "postavshchik"):
            return "unsupported"
        # У Hygiene возвраты выгружаются построчно, у Innowave — документами.
        return "return_docs" if org == "innowave" else "return_lines"
    if has("остатк", "ostatk"):
        # «Остатки …»: деньги, если названы деньги/банк/касса, иначе товары.
        if has("денег", "денежн", "deneg", "denejn", "банк", "bank", "касс", "kass"):
            return "cash_balances"
        return "stock_balances"
    # «Поступление ТОВАРОВ» — закупка у поставщика, а не деньги. Правило стоит
    # раньше денежного «поступления», иначе закупки грузились бы как оплаты.
    if has("поступлен", "postuplen") and has("товар", "tovar"):
        return "purchases"
    # «Дополнительные расходы» — это себестоимость импорта (landed cost),
    # привязанная к ГТД, а не выплата денег. Правило обязано стоять раньше
    # общего «расход», иначе 294 строки по Хайджин уехали бы в денежные
    # расходы и задвоили их.
    if has("дополнительн", "dopolnitel") and has("расход", "rashod"):
        return "unsupported"
    # «Платёжный ордер списание ДС» — банковская операция, а не списание
    # товаров. Правило обязано стоять раньше общего «спис», иначе документ
    # уедет в товарный импортёр, где ждут номенклатуру и количество.
    if has("ордер", "order") and has("списан", "spisan"):
        return "expense"
    # Списание товаров: «ВыгрузкаСпис» и полное «Списание товаров». Правило
    # обязано стоять раньше блока неподдерживаемых — там «списан» до сих пор
    # значилось как вид без импортёра.
    if has("списан", "spisan", "спис", "spis"):
        return "writeoffs"
    # Оприходование товаров — зеркало списания: приход на склад без
    # поставщика (излишки инвентаризации, возврат из эксплуатации).
    if has("оприходован", "oprihodovan"):
        return "stock_receipts"
    # Справочник номенклатуры. Правило обязано стоять раньше блока
    # неподдерживаемых: там «номенклатур» до сих пор значился как вид без
    # импортёра. Товарные выгрузки сюда не попадают — у них в имени вид
    # документа («Реализация», «Поступление»), а не слово «Номенклатура».
    if has("номенклатур", "nomenklatur"):
        return "products"
    # Справочник контрагентов. Как и с номенклатурой, правило обязано стоять
    # раньше блока неподдерживаемых. Документы сюда не попадают: у них в
    # имени вид документа, а слово «Контрагенты» стоит у самого справочника.
    if has("контрагент", "kontragent"):
        return "counterparties"
    # Виды, для которых импортёров пока нет. Ловим по имени осознанно, а не
    # отдаём угадыванию по колонкам: «Оприходование» содержит «приход» и без
    # этого правила уехало бы в денежные поступления. Блок стоит раньше правил
    # поступлений/расходов именно из-за таких пересечений.
    if has("перемещен", "peremeshen", "peremeshch",
           "инвентариз", "inventariz",
           "взаимозач", "vzaimozach",
           "авансов", "avansov",  # авансовый отчёт подотчётника
           "подотчет", "подотчёт", "podotchet",
           "счет на оплату", "счёт на оплату", "schet na oplatu",
           "контрагент", "kontragent",
           "номенклатур", "nomenklatur",
           "оборотно", "oborotno",
           # Новый пакет выгрузки добавил виды, которых в старом не было.
           # Ловим их по имени, потому что sniff_kind по колонкам ошибается:
           # «Движение МБП» с колонками Дата/Сумма/Контрагент он принимает за
           # денежный расход и подмешивает движение малоценки к платежам.
           "гтд", "gtd",
           "журнал проводок", "jurnal provodok", "zhurnal provodok",
           "конвертац", "konvertac",
           "начисление зарплат", "nachislenie zarplat",
           "проблемные документ", "problemnye dokument",
           "ручные операц", "ruchnye operac",
           "движение мбп", "dvijenie mbp", "dvizhenie mbp",
           "эсф", "esf", "счет-фактур", "счёт-фактур", "schet-faktur",
           "бланки счетов", "blanki schetov"):
        return "unsupported"
    # Поступления денег: «Платёжное поручение ВХОДЯЩЕЕ» (банк) и «ПРИХОДНЫЙ
    # кассовый ордер» (касса). Проверяется раньше расходов: слово «платёжное»
    # есть в обоих поручениях, направление решают «входящее»/«исходящее».
    if has("входящ", "vhodyash", "vkhodyash", "приход", "prihod",
           "поступлен", "postuplen"):
        return "receipts"  # банк или касса — решается ниже по слову в имени
    if has("исходящ", "ishodyash", "расход", "rashod",
           "платеж", "платёж", "platej", "poruchenie"):
        return "expense"  # банк или касса — решается ниже по слову в имени
    # Короткое «Пост» (ВыгрузкаПост) — тоже закупки. Правило обязано стоять
    # РАНЬШЕ старого токена «ост»: иначе «…Пост» распознавался как остатки
    # товаров, и закупки летели в чужой импортёр (он, к счастью, отбивался
    # по колонкам, но файл вечно висел в ошибках автосинка).
    if has("пост", "post") and not has("поставщик", "postavshik"):
        return "purchases"

    # --- Продажи ---
    if "реал" in name:
        if org == "innowave":
            return "sales"  # у Innowave каноничен обычный Реал
        return "sales" if "реал2" in name else "dup_sales"

    # --- Возвраты ---
    if "товвозв" in name:
        return "return_lines"
    if "возв" in name:
        # У Innowave нет построчного ТовВозв — значит документный Возв каноничен
        return "return_docs" if org == "innowave" else "dup_returns"

    # --- Остальные типы (по одному файлу) ---
    if "банккасса" in name:
        return "cash_balances"
    if "пписход" in name or "рко" in name:
        return "expense"
    if "банквх" in name or "пко" in name:
        return "receipts"
    if "ост" in name and "прост" not in name:
        return "stock_balances"
    return None


# Названия товарной колонки в разных поколениях выгрузок: старые файлы
# (ВыгрузкаРеал2, ВыгрузкаТовВозв) пишут «НоменклатураНаименование», новый
# пакет «Фирма_УПРАВЛЕНКА_*» — просто «Номенклатура».
_PRODUCT_HEADERS = {"НоменклатураНаименование", "Номенклатура"}


def _headers(content: bytes) -> set[str]:
    """Заголовки колонок из первых 20 строк книги (шапка бывает не в 1-й)."""
    try:
        wb = xlsx.load_workbook(content)
    except Exception:
        return set()
    ws = wb[wb.sheetnames[0]]
    out: set[str] = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 20:
            break
        out |= {str(c).strip() for c in row if c is not None}
    return out


def _probe_columns(content: bytes) -> str:
    """Строка заголовков файла — чтобы понять, из чего писать импортёр.

    Берём среди первых 20 строк ту, где больше всего непустых текстовых
    ячеек: у выгрузок 1С шапка бывает не первой строкой, зато она всегда
    самая «словесная» — ниже идут даты и числа. Сам файл нигде не оседает,
    в журнал попадают только названия колонок."""
    try:
        wb = xlsx.load_workbook(content)
    except Exception:
        return ""
    ws = wb[wb.sheetnames[0]]
    best: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 20:
            break
        cells = [str(c).strip() for c in row
                 if c is not None and str(c).strip()
                 and not isinstance(c, (int, float))]
        if len(cells) > len(best):
            best = cells
    return ", ".join(best[:40])


def _is_line_doc(content: bytes) -> bool:
    """True — выгрузка построчная (по товарам), False — документная (по шапкам).

    Признак строки товара: колонка номенклатуры И количество. Одной
    номенклатуры мало: она встречается и в документных отчётах как справочное
    поле, а количество бывает только у товарных строк."""
    heads = _headers(content)
    if not heads:
        return True  # не смогли прочитать — старое поведение (построчный)
    return bool(heads & _PRODUCT_HEADERS) and "Количество" in heads


def _is_line_sales(content: bytes) -> bool:
    """Оставлено ради читаемости вызова в диспетчере продаж."""
    return _is_line_doc(content)


def sniff_kind(content: bytes) -> str:
    """Определяет тип выгрузки по заголовкам колонок в первых 20 строках."""
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content), data_only=True, read_only=True
        )
    except Exception:
        return "not_excel"
    ws = wb[wb.sheetnames[0]]
    first_rows: list[set] = []
    all_text = ""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 20:
            break
        cells = {str(c).strip() for c in row if c is not None}
        first_rows.append(cells)
        all_text += " " + " ".join(cells)

    # Возвраты выгружаются с теми же колонками, что продажи — отличаем по
    # служебной строке 1С «Запрос: Документ.ВозвратТоваров…» или по слову
    # «Возврат» в шапке. Это спасает выручку от чужих строк.
    is_return_marked = "Возврат" in all_text

    for cells in first_rows:
        if "НоменклатураНаименование" in cells and "Сумма" in cells:
            return "return_lines" if is_return_marked else "sales"
        if {"Дата", "Сумма", "Контрагент", "ВидОперации"} <= cells:
            return "receipts"
        if "Основание" in cells or {"Документ", "Номер"} <= cells:
            return "expense"  # РКО/платёжки — импортёр в разработке
        if "Касса_Банк" in cells and "СуммаОстаток" in cells:
            return "cash_balances"
        if "СуммаОстаток" in cells and "КоличествоОстаток" in cells:
            return "stock_balances"
        # Матричный отчёт остатков: «Номенклатура» + колонка «ИТОГО …»
        if "Номенклатура" in cells and any("ИТОГО" in str(c) for c in cells):
            return "stock_balances"
        if {"Дата", "Сумма", "Валюта", "Контрагент"} <= cells:
            return "return_docs"
    return "unknown"


@router.post("/inbox")
async def inbox(
    file: UploadFile,
    db: Session = Depends(get_db),
    authorization: str | None = Header(default=None),
    fname: str | None = Form(default=None),
    org: str = Form(default=models.DEFAULT_ORG),
    ledger: str = Form(default="upr"),
):
    _require_token(authorization)

    org = models.normalize_org(org)
    content = await file.read()
    # Имя из поля формы (fname) — приходит корректным UTF-8, в отличие от
    # имени в заголовке multipart, где кириллица портится. Заголовок —
    # запасной вариант с попыткой восстановления.
    filename = fname or _recover_filename(file.filename or "file.xlsx")
    # Фирма: сначала по имени файла, папка — запасной вариант. Так одна папка
    # Drive обслуживает все базы, и перекладывать файлы не нужно.
    org = models.normalize_org(org_from_name(filename) or org)
    # Имя файла — первичный сигнал; колонки — запасной для незнакомых имён.
    kind = classify_by_name(filename, org) or sniff_kind(content)
    if kind == "ignore":
        return {"type": "ignore", "status": "skipped", "detail": "Временный файл"}
    # Налоговый контур: файлы из налоговой папки Drive (ledger=nal) или с
    # меткой НАЛ в имени грузятся своими импортёрами в отдельную таблицу —
    # автоматом, как и управленка. Тип определяется по колонкам, поэтому
    # работает даже с черновыми именами без метки.
    if (ledger or "").strip().lower().startswith("nal") or kind == "tax":
        from .tax import import_tax_workbook
        robot = _robot_user(db)
        try:
            result = import_tax_workbook(db, content, filename, robot.id, org)
            return {"type": f"tax_{result['kind']}", "status": "imported", **result}
        except HTTPException as e:
            # Вид, который налоговый импортёр пока не понимает (банк, остатки…)
            # — пропускаем с причиной, не роняя автосинк.
            db.rollback()
            return {"type": "tax_skip", "status": "skipped",
                    "detail": f"Налоговый файл не загружен: {e.detail}"}
    if kind == "unsupported":
        # Раньше такой файл исчезал бесследно, и понять, что 1С присылает
        # оприходования с инвентаризациями, было неоткуда — а именно они
        # объясняют расхождение остатков. Отмечаем приход в журнале: один
        # раз на файл, повторную отправку того же содержимого не дублируем.
        file_hash = hashlib.sha256(content).hexdigest()
        logged = f"[авто:{org}] [не ведём] {filename}"
        # Повтор ловим по паре «имя + содержимое». Одного хэша мало: две
        # разные пустые выгрузки бывают побайтово одинаковыми, и по хэшу
        # второй вид молча слился бы с первым — в отчёте пропал бы целый вид.
        known = (db.query(models.ImportLog)
                 .filter(models.ImportLog.file_hash == file_hash,
                         models.ImportLog.filename == logged)
                 .first())
        if known is None:
            db.add(models.ImportLog(
                filename=logged, user_id=_robot_user(db).id,
                added=0, skipped=0, errors_count=0, file_hash=file_hash,
                columns=_probe_columns(content) or None,
            ))
            db.commit()
        return {
            "type": kind,
            "status": "skipped",
            "detail": "Этот вид выгрузки портал пока не ведёт — файл пропущен "
                      "осознанно и данные не искажает; когда добавим "
                      "поддержку, начнём загружать автоматически",
        }
    if kind in ("dup_sales", "dup_returns"):
        # Дубль-вариант выгрузки (напр. старый Реал при наличии Реал2) —
        # не грузим, чтобы не задваивать данные.
        return {
            "type": kind,
            "status": "skipped",
            "detail": "Дубль-вариант выгрузки — грузится каноничный файл",
        }
    robot = _robot_user(db)

    # Раньше здесь стоял короткий выход «файл уже обработан» по хэшу
    # содержимого. Он мешал переразложить файлы после смены логики
    # маршрутизации, поэтому убран: все импортёры идемпотентны — снапшоты
    # (остатки) грузятся заменой, продажи/оплаты/расходы/возвраты
    # дедуплицируются построчно или заменой периода. Хэш по-прежнему
    # пишем в журнал для аудита.
    file_hash = hashlib.sha256(content).hexdigest()

    auto_name = f"[авто:{org}] {filename}"
    try:
        return _dispatch_import(db, kind, content, auto_name, robot, filename,
                                org, file_hash)
    except HTTPException:
        raise
    except Exception as err:  # noqa: BLE001 — важно назвать место сбоя
        # Ответ уходит в журнал Apps Script, а туда traceback целиком не
        # влезает. Отдаём тип, текст и последний кадр из нашего кода — по ним
        # сразу видно, какой импортёр и какая строка споткнулись.
        tb = traceback.extract_tb(err.__traceback__)
        ours = [f for f in tb if "/app/" in f.filename] or tb
        last = ours[-1] if ours else None
        where = (f"{last.filename.split('/')[-1]}:{last.lineno} в {last.name}()"
                 if last else "неизвестно")
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Не удалось загрузить «{filename}» ({kind}): "
                   f"{type(err).__name__}: {err} — {where}",
        ) from err


def _dispatch_import(db, kind, content, auto_name, robot, filename, org, file_hash):
    """Разбор файла нужным импортёром по определённому типу выгрузки."""
    if kind == "sales":
        # Формат реализации: построчный (есть «НоменклатураНаименование») или
        # документный (Дата/Сумма/Контрагент, как у Innowave).
        if _is_line_sales(content):
            result = import_sales_workbook(db, content, auto_name, robot.id, org=org)
        else:
            result = import_sales_docs_workbook(db, content, auto_name, robot.id, org=org)
    elif kind == "receipts":
        # Касса: старое «ПКО» или слово «касса» в новой схеме имён; иначе банк.
        low = filename.lower()
        rcpt_kind = "cash" if any(
            t in low for t in ("пко", "pko", "касс", "kass")) else "bank"
        result = import_receipts_workbook(
            db, content, auto_name, robot.id, kind=rcpt_kind, org=org
        )
    elif kind == "purchases":
        from .purchases import import_purchases_workbook
        result = import_purchases_workbook(db, content, auto_name, robot.id, org=org)
    elif kind == "writeoffs":
        from .writeoffs import import_writeoffs_workbook
        result = import_writeoffs_workbook(db, content, auto_name, robot.id, org=org)
    elif kind == "stock_receipts":
        from .stock_receipts import import_stock_receipts_workbook
        result = import_stock_receipts_workbook(db, content, auto_name,
                                                robot.id, org=org)
    elif kind == "products":
        from .products import import_products_workbook
        result = import_products_workbook(db, content, auto_name, robot.id,
                                          org=org)
    elif kind == "counterparties":
        from .counterparties import import_counterparties_workbook
        result = import_counterparties_workbook(db, content, auto_name,
                                                robot.id, org=org)
    elif kind == "return_docs" and not _is_line_doc(content):
        result = import_returns_workbook(db, content, auto_name, robot.id, org=org)
    elif kind == "cash_balances":
        result = import_cash_balances_workbook(db, content, auto_name, robot.id, org=org)
    elif kind == "stock_balances":
        result = import_stock_balances_workbook(db, content, auto_name, robot.id, org=org)
    elif kind == "expense":
        # Касса: старое «РКО» или слово «касса» в новой схеме имён; иначе банк.
        low = filename.lower()
        exp_kind = "cash" if any(
            t in low for t in ("рко", "rko", "касс", "kass")) else "bank"
        result = import_expenses_workbook(db, content, auto_name, robot.id, exp_kind, org=org)
    elif kind in ("return_lines", "return_docs"):
        # Построчные возвраты: очистка продаж + запись сумм по клиентам.
        # Сюда же попадает return_docs, если файл оказался построчным: формат
        # зависит от выгрузки, а не от фирмы. У Innowave раньше был только
        # документный формат, и классификатор решал по названию фирмы — с
        # новым пакетом «Фирма_УПРАВЛЕНКА_*» обе фирмы шлют построчный.
        result = import_return_lines_workbook(db, content, auto_name, robot.id, org=org)
        log = db.query(models.ImportLog).order_by(models.ImportLog.id.desc()).first()
        if log and log.file_hash is None:
            log.file_hash = file_hash
            db.commit()
        return {"type": kind, "status": "imported", **result}
    else:
        db.add(models.ImportLog(
            filename=f"[авто, не распознан] {filename}",
            user_id=robot.id,
            added=0,
            skipped=0,
            errors_count=0,
        ))
        db.commit()
        return {
            "type": kind,
            "status": "skipped",
            "detail": "Формат пока не поддерживается — файл записан в журнал",
        }

    # Проставляем хэш файла в свежую запись журнала
    log = (
        db.query(models.ImportLog)
        .order_by(models.ImportLog.id.desc())
        .first()
    )
    if log and log.file_hash is None:
        log.file_hash = file_hash
        db.commit()

    return {"type": kind, "status": "imported", **result}


admin_only = require_roles(models.Role.admin)


@router.get("/skipped-kinds")
def skipped_kinds(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Виды выгрузок, которые 1С присылает, а портал не грузит.

    Отчёт отвечает на вопрос, который иначе некуда задать: почему расчётный
    остаток не сходится с 1С. Часть присланных документов двигает склад —
    оприходования, инвентаризации, корректировки, возвраты поставщику, — но
    импортёров для них нет, файлы пропускаются, и в математике этих движений
    просто не существует. Пока такой документ не загружен, расхождение по
    товару объяснить нечем: в наших движениях его нет ни в каком виде."""
    logs = (db.query(models.ImportLog)
            .filter(models.ImportLog.filename.like("%[не ведём]%"))
            .order_by(models.ImportLog.created_at.desc())
            .all())
    groups: dict[str, dict] = {}
    for l in logs:
        label, moves, note = unsupported_kind(l.filename)
        g = groups.setdefault(label, {
            "kind": label, "moves_stock": moves, "note": note, "files": 0,
            "last_at": None, "last_file": None, "columns": None,
        })
        g["files"] += 1
        if g["last_at"] is None:  # выборка уже отсортирована по убыванию
            g["last_at"] = l.created_at.isoformat()
            g["last_file"] = l.filename.split("] ")[-1]
            g["columns"] = l.columns
    rows = sorted(groups.values(),
                  key=lambda g: (not g["moves_stock"], -g["files"]))
    return {
        "rows": rows,
        "files": len(logs),
        "moving_files": sum(g["files"] for g in rows if g["moves_stock"]),
    }


@router.post("/reset")
def reset_imported_data(
    db: Session = Depends(get_db),
    _: models.User = Depends(admin_only),
):
    """Полная очистка импортированных из 1С данных для чистого переимпорта.

    Удаляет только то, что грузится из 1С. Ручные данные портала —
    пользователи, платежи календаря, планы агентов, сопоставления имён,
    принятые нарушения — сохраняются.
    """
    cleared = {}
    for model in (
        models.Sale,
        models.Receipt,
        models.Expense,
        models.ReturnDoc,
        models.CashBalance,
        models.StockBalance,
        models.ImportLog,
    ):
        cleared[model.__tablename__] = db.query(model).delete()
    db.commit()
    return {"status": "reset", "cleared": cleared}
