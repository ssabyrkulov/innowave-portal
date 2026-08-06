import hashlib
import io
import re
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from fastapi import (
    APIRouter,
    Depends,
    Form,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import get_current_user, require_roles

router = APIRouter(prefix="/receipts", tags=["receipts"])

can_import = require_roles(models.Role.admin, models.Role.accountant)
admin_only = require_roles(models.Role.admin)

# Курсы к сому по умолчанию для пересчёта валютных поступлений.
# Курс хранится в каждой строке и правится вручную в интерфейсе.
DEFAULT_RATES = {"KGS": 1.0, "USD": 87.0, "EUR": 95.0, "RUB": 1.1, "KZT": 0.17}

HEADERS = {"Дата": "date", "Сумма": "amount", "Валюта": "currency",
           "Контрагент": "payer", "ВидОперации": "operation",
           # Необязательная колонка обновлённых выгрузок 1С: GUID документа.
           "ДокументGUID": "doc_guid"}

# Дебиторку формируют только оплаты покупателей.
CUSTOMER_PAYMENT_PREFIX = "Оплата от покупателя"


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.split()[0], "%d.%m.%Y").date()
        except ValueError:
            return None
    return None


def _hash(d: dict, occurrence: int, org: str = models.DEFAULT_ORG) -> str:
    # В ключе — организация + исходная дата с точным временем (в 1С это разные
    # документы, даже если сумма и день совпадают) плюс порядковый номер
    # повтора на случай полностью идентичных строк выгрузки.
    key = "|".join(str(d[f]) for f in ("src_dt", "amount", "currency", "payer", "operation"))
    return hashlib.sha256(f"{org}|{key}#{occurrence}".encode()).hexdigest()


def _normalize(name: str) -> str:
    return re.sub(r"\s+", " ", name.lower().replace("ё", "е")).strip()


@router.post("/import")
async def import_receipts(
    file: UploadFile,
    replace_period: bool = Form(default=False),
    org: str = Form(default=models.DEFAULT_ORG),
    db: Session = Depends(get_db),
    current: models.User = Depends(can_import),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Ожидается файл Excel (.xlsx)")
    content = await file.read()
    return import_receipts_workbook(
        db, content, file.filename, current.id, replace_period, org=org
    )


def import_receipts_workbook(
    db: Session,
    content: bytes,
    filename: str,
    user_id: int,
    replace_period: bool = False,
    kind: str = "bank",
    org: str = models.DEFAULT_ORG,
) -> dict:
    """Импорт поступлений из байтов Excel (веб-загрузка и автоприём).

    kind — источник денег: 'bank' (ВыгрузкаБанкВх) или 'cash' (ВыгрузкаПКО).
    В row_hash не входит, поэтому на дедупликацию не влияет.
    org — организация (в хэш входит, чтобы одинаковые строки разных фирм не
    считались дублями).
    """
    org = models.normalize_org(org)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось открыть файл Excel")

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)

    header_idx, columns, buffered = None, {}, []
    for i, row in enumerate(rows_iter):
        buffered.append(row)
        if i >= 20:
            break
        matched = {
            j: HEADERS[str(c).strip()]
            for j, c in enumerate(row)
            if c is not None and str(c).strip() in HEADERS
        }
        if {"date", "amount", "payer"} <= set(matched.values()):
            header_idx, columns = i, matched
            break
    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Не найдены колонки поступлений (Дата, Сумма, Валюта, "
                   "Контрагент, ВидОперации)",
        )

    parsed_rows, errors = [], []

    def process(row, line_no):
        data = {f: (row[j] if j < len(row) else None) for j, f in columns.items()}
        if all(v is None for v in data.values()):
            return
        dt = _parse_date(data.get("date"))
        try:
            amount = float(str(data.get("amount")).replace(" ", "").replace(",", "."))
        except (TypeError, ValueError):
            amount = None
        # Контрагент бывает пуст у служебных операций (снятие наличных,
        # прочий приход) — такие строки сохраняем, в дебиторку они не идут.
        payer = str(data.get("payer") or "").strip() or "Не указан"
        if not dt or amount is None:
            if len(errors) < 20:
                errors.append(f"Строка {line_no}: нет даты или суммы")
            return
        currency = (str(data.get("currency") or "KGS").strip() or "KGS")[:3]
        rate = DEFAULT_RATES.get(currency, 1.0)
        parsed_rows.append({
            "src_dt": str(data.get("date")),
            "date": dt,
            "amount": amount,
            "currency": currency,
            "rate": rate,
            "amount_kgs": round(amount * rate, 2),
            "payer": payer,
            "operation": str(data.get("operation") or "").strip() or "Не указан",
            "doc_guid": str(data.get("doc_guid") or "").strip() or None,
        })

    line_no = header_idx + 1
    for line_no, row in enumerate(buffered[header_idx + 1:], start=header_idx + 2):
        process(row, line_no)
    for line_no, row in enumerate(rows_iter, start=line_no + 1):
        process(row, line_no)

    replaced = 0
    if replace_period and parsed_rows:
        dmin = min(p["date"] for p in parsed_rows)
        dmax = max(p["date"] for p in parsed_rows)
        replaced = (
            db.query(models.Receipt)
            .filter(models.Receipt.organization == org,
                    models.Receipt.date >= dmin, models.Receipt.date <= dmax)
            .delete(synchronize_session=False)
        )
        db.flush()

    existing = {h for (h,) in db.query(models.Receipt.row_hash).all()}
    seen, added, skipped = set(), 0, 0
    occurrences: dict[str, int] = defaultdict(int)
    for p in parsed_rows:
        base = "|".join(str(p[f]) for f in ("src_dt", "amount", "currency", "payer", "operation"))
        occurrences[base] += 1
        h = _hash(p, occurrences[base], org)
        p = {k: v for k, v in p.items() if k != "src_dt"}
        if h in existing or h in seen:
            skipped += 1
            continue
        seen.add(h)
        db.add(models.Receipt(**p, kind=kind, organization=org, row_hash=h))
        added += 1

    db.add(models.ImportLog(
        filename=f"[оплаты] {filename}",
        user_id=user_id,
        added=added,
        skipped=skipped,
        errors_count=len(errors),
        replace_period=replace_period,
    ))
    db.commit()
    return {
        "added": added,
        "skipped_duplicates": skipped,
        "replaced_rows": replaced,
        "errors": errors,
        "total_in_db": db.query(models.Receipt).count(),
    }


class RateUpdate(BaseModel):
    rate: float = Field(gt=0)


@router.patch("/{receipt_id}/rate")
def update_rate(
    receipt_id: int,
    payload: RateUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(can_import),
):
    r = db.get(models.Receipt, receipt_id)
    if not r:
        raise HTTPException(status_code=404, detail="Поступление не найдено")
    r.rate = payload.rate
    r.amount_kgs = round(float(r.amount) * payload.rate, 2)
    db.commit()
    return {"id": r.id, "rate": float(r.rate), "amount_kgs": float(r.amount_kgs)}


@router.get("")
def list_receipts(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    operation: str | None = Query(default=None),
    limit: int = Query(default=300, le=1000),
    org: str = Query(default="all"),
):
    q = models.org_scope(db.query(models.Receipt), models.Receipt, org).order_by(models.Receipt.date.desc())
    if operation:
        q = q.filter(models.Receipt.operation == operation)
    return [
        {
            "id": r.id,
            "date": r.date.isoformat(),
            "amount": float(r.amount),
            "currency": r.currency,
            "rate": float(r.rate),
            "amount_kgs": float(r.amount_kgs),
            "payer": r.payer,
            "operation": r.operation,
            "kind": r.kind or "bank",
        }
        for r in q.limit(limit).all()
    ]


class AliasCreate(BaseModel):
    payer: str
    client: str


@router.post("/alias", status_code=status.HTTP_201_CREATED)
def create_alias(
    payload: AliasCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(can_import),
):
    existing = db.query(models.ClientAlias).filter_by(payer=payload.payer).first()
    if existing:
        existing.client = payload.client
    else:
        db.add(models.ClientAlias(payer=payload.payer, client=payload.client))
    db.commit()
    return {"status": "ok"}


@router.delete("/alias/{alias_payer:path}", status_code=status.HTTP_204_NO_CONTENT)
def delete_alias(
    alias_payer: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(can_import),
):
    a = db.query(models.ClientAlias).filter_by(payer=alias_payer).first()
    if not a:
        raise HTTPException(status_code=404, detail="Сопоставление не найдено")
    db.delete(a)
    db.commit()


FLAG_KINDS = {"bad_debt", "disputed"}


class ClientFlagCreate(BaseModel):
    client: str
    kind: str = "bad_debt"
    note: str | None = None


@router.get("/flags")
def list_flags(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    rows = (
        db.query(models.BadDebtClient)
        .order_by(models.BadDebtClient.created_at.desc())
        .all()
    )
    return [
        {
            "client": r.client,
            "kind": r.kind or "bad_debt",
            "note": r.note,
            "created_at": r.created_at.isoformat(),
            "created_by": r.creator.full_name if r.creator else None,
        }
        for r in rows
    ]


@router.post("/flags", status_code=status.HTTP_201_CREATED)
def set_flag(
    payload: ClientFlagCreate,
    db: Session = Depends(get_db),
    current: models.User = Depends(can_import),
):
    client = payload.client.strip()
    if not client:
        raise HTTPException(status_code=400, detail="Не указан контрагент")
    kind = payload.kind if payload.kind in FLAG_KINDS else "bad_debt"
    note = (payload.note or "").strip() or None
    existing = db.query(models.BadDebtClient).filter_by(client=client).first()
    if existing:
        existing.kind = kind
        existing.note = note
    else:
        db.add(models.BadDebtClient(
            client=client, kind=kind, note=note, created_by=current.id,
        ))
    db.commit()
    return {"status": "ok", "client": client, "kind": kind}


@router.delete("/flags/{client:path}", status_code=status.HTTP_204_NO_CONTENT)
def remove_flag(
    client: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(can_import),
):
    row = db.query(models.BadDebtClient).filter_by(client=client).first()
    if not row:
        raise HTTPException(status_code=404, detail="Клиент не помечен")
    db.delete(row)
    db.commit()


@router.get("/client-detail")
def client_detail(
    client: str = Query(..., description="Имя контрагента из продаж"),
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    org: str = Query(default="all"),
):
    """Карточка контрагента: отгрузки, возвраты и оплаты по датам + итоги.

    Оплаты и возвраты сопоставляются с клиентом так же, как в дебиторке —
    через ручные сопоставления имён (алиасы) и нормализацию названия.
    """
    target = client.strip()
    norm_target = _normalize(target)
    aliases = {a.payer: a.client for a in db.query(models.ClientAlias).all()}

    def resolves(name: str) -> bool:
        c = aliases.get(name)
        if c is not None:
            return c == target
        return name == target or _normalize(name) == norm_target

    # --- Отгрузки: сводим строки продаж в документы ---
    sales = models.org_scope(
        db.query(models.Sale).filter(models.Sale.client == target), models.Sale, org
    ).all()
    by_doc: dict = {}
    loose: list[dict] = []
    for s in sales:
        line_amt = float(s.amount) * (1 - float(s.discount_pct or 0) / 100)
        if s.doc_number:
            key = (s.doc_number, s.date)
            d = by_doc.get(key)
            if d is None:
                d = by_doc[key] = {
                    "date": s.date.isoformat(),
                    "doc_number": s.doc_number,
                    "amount": 0.0,
                    "doc_total": float(s.doc_total) if s.doc_total is not None else None,
                    "positions": 0,
                    # Склад отгрузки — по нему в сверке видно, к какой фирме
                    # относится документ; в SalesDoc склад свой, и расхождение
                    # часто объясняется именно разными складами.
                    "warehouses": [],
                }
            d["positions"] += 1
            d["amount"] += line_amt
            if s.warehouse and s.warehouse not in d["warehouses"]:
                d["warehouses"].append(s.warehouse)
            if s.doc_total is not None:
                d["doc_total"] = float(s.doc_total)
        else:
            loose.append({
                "date": s.date.isoformat(),
                "doc_number": None,
                "amount": round(line_amt, 2),
                "doc_total": None,
                "positions": 1,
                "warehouses": [s.warehouse] if s.warehouse else [],
            })

    shipments = []
    shipped = 0.0
    for d in list(by_doc.values()) + loose:
        amt = d["doc_total"] if d["doc_total"] is not None else d["amount"]
        amt = round(float(amt), 2)
        shipped += amt
        shipments.append({
            "date": d["date"],
            "doc_number": d["doc_number"],
            "amount": amt,
            "positions": d["positions"],
            "warehouse": ", ".join(d.get("warehouses") or []) or None,
        })
    shipments.sort(key=lambda x: x["date"], reverse=True)

    # --- Возвраты ---
    returns = []
    returned = 0.0
    for rd in models.org_scope(db.query(models.ReturnDoc), models.ReturnDoc, org).all():
        if resolves(rd.client):
            amt = float(rd.amount)
            returned += amt
            returns.append({"date": rd.date.isoformat(), "amount": round(amt, 2)})
    returns.sort(key=lambda x: x["date"], reverse=True)

    # --- Оплаты ---
    payments = []
    paid = 0.0
    for r in models.org_scope(db.query(models.Receipt), models.Receipt, org).all():
        if not r.operation.startswith(CUSTOMER_PAYMENT_PREFIX):
            continue
        if not resolves(r.payer):
            continue
        amt = float(r.amount_kgs)
        paid += amt
        payments.append({
            "date": r.date.isoformat(),
            "amount": float(r.amount),
            "amount_kgs": round(amt, 2),
            "currency": r.currency,
            "operation": r.operation,
            "kind": r.kind or "bank",
        })
    payments.sort(key=lambda x: x["date"], reverse=True)

    return {
        "client": target,
        "totals": {
            "shipped": round(shipped, 2),
            "returned": round(returned, 2),
            "paid": round(paid, 2),
            "debt": round(shipped - returned - paid, 2),
        },
        "shipments": shipments,
        "returns": returns,
        "payments": payments,
    }


@router.get("/receivables")
def receivables(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    org: str = Query(default="all"),
):
    """Дебиторка: отгружено − возвраты − оплачено по каждому клиенту."""
    from ..services import salesdoc_mirror

    sales = models.org_scope(db.query(models.Sale), models.Sale, org).all()
    receipts = models.org_scope(db.query(models.Receipt), models.Receipt, org).all()
    return_docs = models.org_scope(db.query(models.ReturnDoc), models.ReturnDoc, org).all()
    aliases = {a.payer: a.client for a in db.query(models.ClientAlias).all()}
    flag_rows = db.query(models.BadDebtClient).all()
    flags = {b.client: (b.kind or "bad_debt") for b in flag_rows}
    flag_notes = {b.client: b.note for b in flag_rows if b.note}

    # Отгрузки по клиентам: итог документа (после скидок); документы без
    # номера учитываем по строкам со скидкой.
    shipped: dict[str, float] = defaultdict(float)
    last_shipment: dict[str, date] = {}
    client_org: dict[str, str] = {}
    seen_docs: set = set()
    for s in sales:
        client = s.client
        client_org.setdefault(client, s.organization)
        if s.doc_number:
            key = (s.doc_number, s.date, client)
            if key in seen_docs:
                continue
            if s.doc_total is not None:
                seen_docs.add(key)
                shipped[client] += float(s.doc_total)
            else:
                shipped[client] += float(s.amount) * (
                    1 - float(s.discount_pct or 0) / 100
                )
        else:
            shipped[client] += float(s.amount) * (
                1 - float(s.discount_pct or 0) / 100
            )
        if client not in last_shipment or s.date > last_shipment[client]:
            last_shipment[client] = s.date

    norm_clients = {_normalize(c): c for c in shipped}

    # Агент точки. Долг взыскивает живой человек, поэтому колонка отвечает не
    # на вопрос «кто когда-то продал», а «кому сегодня ставить задачу»:
    # действующий агент из SalesDoc, а если точки там нет — агент последней
    # реализации 1С, помеченный уволенным, если в справочнике он неактивен.
    try:
        idx = salesdoc_mirror.agent_by_client_1c(db)
    except Exception:  # noqa: BLE001 — дебиторка обязана открыться и без SalesDoc
        idx = {"by_client": {}, "agents": [], "active_names": []}
    agent_of = idx["by_client"]
    _ag_key = salesdoc_mirror.match_key
    known_agents = {_ag_key(a["name"]): a["active"] for a in idx["agents"]}

    last_1c_agent: dict[str, tuple[date, str]] = {}
    for s in sales:
        name = s.agent or s.responsible
        if not name:
            continue
        prev = last_1c_agent.get(s.client)
        if prev is None or s.date > prev[0]:
            last_1c_agent[s.client] = (s.date, name)

    # Возвраты уменьшают долг клиента (сопоставление имён — как у оплат)
    returned: dict[str, float] = defaultdict(float)
    for rd in return_docs:
        client = aliases.get(rd.client)
        if client is None:
            client = rd.client if rd.client in shipped else norm_clients.get(_normalize(rd.client))
        returned[client or rd.client] += float(rd.amount)

    paid: dict[str, float] = defaultdict(float)
    last_payment: dict[str, date] = {}
    unmatched: dict[str, dict] = {}
    # Банк и касса раздельно, и в каждом — сколько строк дошло до дебиторки.
    # Две причины, по которым касса может не влиять на долг, выглядят снаружи
    # одинаково («долг завышен»), а лечатся по-разному: файла ПКО нет вовсе
    # либо он грузится, но с видом операции, который к покупателям отношения
    # не имеет (розничная выручка, снятие в банке, возврат подотчётника).
    # Поэтому считаем обе цифры и отдаём наружу, а не пишем вывод в вёрстке.
    def _kind_of(r) -> str:
        return r.kind or "bank"

    by_kind: dict[str, dict] = {
        "bank": {"paid": 0.0, "count": 0, "rows": 0},
        "cash": {"paid": 0.0, "count": 0, "rows": 0},
    }
    other_ops: dict[str, dict] = {}
    for r in receipts:
        k = by_kind.setdefault(_kind_of(r), {"paid": 0.0, "count": 0, "rows": 0})
        k["rows"] += 1
        if not r.operation.startswith(CUSTOMER_PAYMENT_PREFIX):
            key = (_kind_of(r), r.operation)
            o = other_ops.setdefault(key, {
                "kind": key[0], "operation": r.operation,
                "amount": 0.0, "count": 0,
            })
            o["amount"] += float(r.amount_kgs)
            o["count"] += 1
            continue
        k["paid"] += float(r.amount_kgs)
        k["count"] += 1
        client = aliases.get(r.payer)
        if client is None:
            if r.payer in shipped:
                client = r.payer
            else:
                client = norm_clients.get(_normalize(r.payer))
        if client is None:
            u = unmatched.setdefault(r.payer, {"payer": r.payer, "paid": 0.0, "count": 0})
            u["paid"] += float(r.amount_kgs)
            u["count"] += 1
            continue
        paid[client] += float(r.amount_kgs)
        if client not in last_payment or r.date > last_payment[client]:
            last_payment[client] = r.date

    clients = []
    for client in set(shipped) | set(paid) | set(returned):
        sh = shipped.get(client, 0.0)
        pd = paid.get(client, 0.0)
        ret = returned.get(client, 0.0)
        a = agent_of.get(client)
        if a is None and client in last_1c_agent:
            when, name = last_1c_agent[client]
            a = {"agent": name,
                 # Неизвестного справочнику агента не объявляем уволенным:
                 # в 1С встречаются и менеджеры, которых в SalesDoc нет вовсе.
                 "agent_active": known_agents.get(_ag_key(name), True),
                 "agent_source": "1С", "agent_at": when.isoformat(),
                 "agent_days": None}
        clients.append({
            "agent": a and a["agent"],
            "agent_active": a["agent_active"] if a else None,
            "agent_source": a and a["agent_source"],
            "agent_at": a and a["agent_at"],
            "agent_days": a and a.get("agent_days"),
            "client": client,
            "shipped": round(sh, 2),
            "returned": round(ret, 2),
            "paid": round(pd, 2),
            "debt": round(sh - ret - pd, 2),
            "organization": client_org.get(client),
            "last_shipment": last_shipment.get(client) and last_shipment[client].isoformat(),
            "last_payment": last_payment.get(client) and last_payment[client].isoformat(),
        })
    clients.sort(key=lambda c: -c["debt"])

    by_agent: dict[str, dict] = {}
    for c in clients:
        if not c["agent"]:
            continue
        a = by_agent.setdefault(c["agent"], {
            "name": c["agent"], "active": c["agent_active"],
            "clients": 0, "debt": 0.0,
        })
        a["clients"] += 1
        a["debt"] += c["debt"]
    agent_list = sorted(
        ({**a, "debt": round(a["debt"], 2)} for a in by_agent.values()),
        key=lambda a: -a["debt"],
    )

    total_shipped = sum(shipped.values())
    total_paid = sum(paid.values())
    total_returned = sum(returned.values())
    return {
        "total_shipped": round(total_shipped, 2),
        "total_returned": round(total_returned, 2),
        "total_paid": round(total_paid, 2),
        "total_debt": round(total_shipped - total_returned - total_paid, 2),
        "clients": clients,
        "unmatched": sorted(unmatched.values(), key=lambda u: -u["paid"]),
        "sales_clients": sorted(shipped.keys()),
        "flags": flags,
        "flag_notes": flag_notes,
        "has_receipts": len(receipts) > 0,
        "payment_kinds": {
            k: {"paid": round(v["paid"], 2), "count": v["count"], "rows": v["rows"]}
            for k, v in by_kind.items()
        },
        # Поступления, которые в дебиторку не идут: это не ошибка (розничная
        # выручка и снятие наличных долг покупателя не гасят), но видеть их
        # надо — иначе «оплата в кассе есть, а долг не уменьшился» выглядит
        # как сбой портала.
        "other_operations": sorted(
            ({**o, "amount": round(o["amount"], 2)} for o in other_ops.values()),
            key=lambda o: -o["amount"],
        ),
        # Для фильтра: только те агенты, за кем реально числится хоть одна
        # точка — справочник целиком в выпадающем списке не нужен.
        "agents": agent_list,
    }
