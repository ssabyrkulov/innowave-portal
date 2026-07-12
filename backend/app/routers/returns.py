"""Возвраты товаров от покупателей (документный уровень, файл ВыгрузкаВозв:
Дата / Сумма / Валюта / Контрагент). Уменьшают долг клиента в дебиторке."""

import hashlib
import io
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import get_current_user

router = APIRouter(prefix="/returns", tags=["returns"])

HEADERS = {"Дата": "date", "Сумма": "amount", "Валюта": "currency",
           "Контрагент": "client"}


def import_return_lines_workbook(
    db: Session, content: bytes, filename: str, user_id: int
) -> dict:
    """Возвраты из построчного файла ВыгрузкаТовВозв (формат продаж).

    Делает две вещи:
    1. Самолечение: удаляет из продаж строки, если они когда-то по ошибке
       туда попали (совпадение по row_hash).
    2. Записывает суммы возвратов по клиентам (заменой целиком). Сумма
       берётся по «СуммаДокумента» возврата (итог документа, как у продаж) —
       так значения совпадают с выгрузкой Возв и Excel. Если у документа нет
       итога — суммируем строки со скидкой.
    """
    from .sales import _row_hash, parse_sales_workbook

    parsed, errors = parse_sales_workbook(content)

    # 1. Очистка продаж от возвратных строк
    hashes = [_row_hash(p) for p in parsed]
    removed = 0
    if hashes:
        removed = (
            db.query(models.Sale)
            .filter(models.Sale.row_hash.in_(hashes))
            .delete(synchronize_session=False)
        )

    # 2. Возвраты по документам: один ReturnDoc на документ по СуммаДокумента.
    db.query(models.ReturnDoc).delete()
    db.flush()
    seen_docs: set = set()
    fallback: dict[tuple, dict] = {}  # документы без номера/итога — по строкам
    docs: list[dict] = []
    for p in parsed:
        client = p["client"]
        if p.get("doc_number") and p.get("doc_total") is not None:
            key = (p["doc_number"], p["date"], client)
            if key in seen_docs:
                continue
            seen_docs.add(key)
            docs.append({
                "date": p["date"], "client": client,
                "amount": float(p["doc_total"]),
                "currency": p.get("currency") or "KGS",
            })
        else:
            # нет итога документа — копим сумму строк (со скидкой) по клиенту/дате
            fkey = (p["date"], client)
            f = fallback.setdefault(fkey, {
                "date": p["date"], "client": client, "amount": 0.0,
                "currency": p.get("currency") or "KGS",
            })
            f["amount"] += float(p["amount"]) * (1 - float(p.get("discount_pct") or 0) / 100)
    docs.extend(fallback.values())

    added = 0
    for i, d in enumerate(docs):
        h = hashlib.sha256(
            f"{d['date']}|{d['client']}|{d['amount']}|{i}".encode()
        ).hexdigest()
        db.add(models.ReturnDoc(
            date=d["date"], amount=round(d["amount"], 2),
            currency=d["currency"], client=d["client"], row_hash=h,
        ))
        added += 1

    db.add(models.ImportLog(
        filename=f"[возвраты, очистка продаж −{removed}] {filename}",
        user_id=user_id, added=added, skipped=0, errors_count=len(errors),
    ))
    db.commit()
    return {
        "added_returns": added,
        "removed_from_sales": removed,
        "total_returns_in_db": db.query(models.ReturnDoc).count(),
    }


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.split()[0], "%d.%m.%Y").date()
        except ValueError:
            return None
    return None


def import_returns_workbook(
    db: Session,
    content: bytes,
    filename: str,
    user_id: int,
) -> dict:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось открыть файл Excel")

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)

    header_idx, columns, buffered = None, {}, []
    for i, row in enumerate(rows_iter):
        buffered.append(row)
        if i >= 20:
            break
        matched = {
            j: HEADERS[str(c).strip()]
            for j, c in enumerate(row)
            if c is not None and str(c).strip() in HEADERS
        }
        if {"date", "amount", "client"} <= set(matched.values()):
            header_idx, columns = i, matched
            break
    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Не найдены колонки возвратов (Дата, Сумма, Валюта, Контрагент)",
        )

    parsed, errors = [], []

    def process(row, line_no):
        data = {f: (row[j] if j < len(row) else None) for j, f in columns.items()}
        if all(v is None for v in data.values()):
            return
        dt = _parse_date(data.get("date"))
        try:
            amount = float(str(data.get("amount")).replace(" ", "").replace(",", "."))
        except (TypeError, ValueError):
            amount = None
        client = str(data.get("client") or "").strip()
        if not dt or amount is None or not client:
            if len(errors) < 20:
                errors.append(f"Строка {line_no}: нет даты, суммы или контрагента")
            return
        parsed.append({
            "src_dt": str(data.get("date")),
            "date": dt,
            "amount": amount,
            "currency": (str(data.get("currency") or "KGS").strip() or "KGS")[:3],
            "client": client,
        })

    line_no = header_idx + 1
    for line_no, row in enumerate(buffered[header_idx + 1:], start=header_idx + 2):
        process(row, line_no)
    for line_no, row in enumerate(rows_iter, start=line_no + 1):
        process(row, line_no)

    # Возвраты — полная выгрузка за всю историю, поэтому грузим «заменой
    # целиком»: это самоисцеляет таблицу, если в неё раньше по ошибке попал
    # чужой файл (напр. исходящие платежи).
    db.query(models.ReturnDoc).delete()
    db.flush()
    existing: set[str] = set()
    seen: set[str] = set()
    occurrences: dict[str, int] = defaultdict(int)
    added = skipped = 0
    for p in parsed:
        base = "|".join(str(p[f]) for f in ("src_dt", "amount", "currency", "client"))
        occurrences[base] += 1
        h = hashlib.sha256(f"{base}#{occurrences[base]}".encode()).hexdigest()
        p = {k: v for k, v in p.items() if k != "src_dt"}
        if h in existing or h in seen:
            skipped += 1
            continue
        seen.add(h)
        db.add(models.ReturnDoc(**p, row_hash=h))
        added += 1

    db.add(models.ImportLog(
        filename=f"[возвраты] {filename}",
        user_id=user_id,
        added=added,
        skipped=skipped,
        errors_count=len(errors),
    ))
    db.commit()
    return {
        "added": added,
        "skipped_duplicates": skipped,
        "errors": errors,
        "total_in_db": db.query(models.ReturnDoc).count(),
    }


@router.get("")
def list_returns(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    limit: int = Query(default=300, le=1000),
):
    rows = (
        db.query(models.ReturnDoc)
        .order_by(models.ReturnDoc.date.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id": r.id,
            "date": r.date.isoformat(),
            "amount": float(r.amount),
            "currency": r.currency,
            "client": r.client,
        }
        for r in rows
    ]
