"""Налоговый контур (черновик): выгрузки из налоговой базы 1С (ред. 1.7).

Пока Эрмек не довёл выгрузку до конца (нет метки НАЛ в именах, нет банка и
остатков), файлы грузятся вручную на отдельной странице и живут в своей
таблице — с управленческими данными не пересекаются вообще. Когда формат
устаканится, эти же импортёры подключатся к автоприёму.
"""

import io
import zipfile
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import get_current_user, require_roles

router = APIRouter(prefix="/tax", tags=["tax"])

can_edit = require_roles(models.Role.admin, models.Role.accountant)


def _load_wb(content: bytes):
    """openpyxl с починкой архива 1С: ред. 1.7 пишет SharedStrings.xml с
    большой буквы, а openpyxl ищет строчную — без переупаковки файл не
    открывается вовсе."""
    try:
        return openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except KeyError:
        zin = zipfile.ZipFile(io.BytesIO(content))
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for it in zin.infolist():
                name = it.filename
                if name.lower().endswith("sharedstrings.xml"):
                    name = "xl/sharedStrings.xml"
                zout.writestr(name, zin.read(it.filename))
        return openpyxl.load_workbook(io.BytesIO(buf.getvalue()), data_only=True, read_only=True)


def _rows(content: bytes) -> list[list]:
    wb = _load_wb(content)
    ws = wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _day(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return datetime.strptime(value.split()[0], "%d.%m.%Y").date()
        except ValueError:
            return None
    return None


def _num(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("\xa0", "").replace(" ", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _header_map(rows: list[list]) -> tuple[int, dict]:
    """Строка заголовков и карта «имя колонки → индекс» (первые 5 строк)."""
    for i, row in enumerate(rows[:5]):
        names = {str(c).strip(): j for j, c in enumerate(row) if c}
        if "Дата" in names and ("Сумма" in names or "СуммаДокумента" in names):
            return i, names
    raise HTTPException(status_code=400, detail="Не нашёл строку заголовков (нужны «Дата» и «Сумма»)")


def _detect_kind(names: dict, filename: str) -> str:
    """Тип файла по заголовкам; имя — только чтобы отличить ПКО от РКО,
    когда колонка вида операции называется одинаково."""
    if "НоменклатураНаименование" in names:
        return "sale"
    if "ВидОперации" in names:
        return "cash_in"      # так называется колонка в ПКО
    if "Основание" in names:
        return "cash_out"     # а так — в РКО
    low = (filename or "").lower()
    if any(t in low for t in ("пко", "pko", "приход", "prihod", "вход", "vhod")):
        return "cash_in"
    if any(t in low for t in ("рко", "rko", "расход", "rashod", "исход", "ishod")):
        return "cash_out"
    return "return"  # Дата/Сумма/Валюта/Контрагент — документные возвраты


def _parse(content: bytes, filename: str) -> tuple[str, list[dict]]:
    rows = _rows(content)
    hi, names = _header_map(rows)
    kind = _detect_kind(names, filename)

    def cell(row, name):
        j = names.get(name)
        return row[j] if j is not None and j < len(row) else None

    out: list[dict] = []
    for row in rows[hi + 1:]:
        d = _day(cell(row, "Дата"))
        if d is None:
            continue
        if kind == "sale":
            amount = _num(cell(row, "Сумма"))
            if amount is None:
                continue
            out.append({
                "kind": kind, "date": d,
                "counterparty": str(cell(row, "КонтрагентНаименование") or "").strip() or None,
                "amount": amount,
                "currency": str(cell(row, "ВалютаДокументаНаименование") or "KGS").strip() or "KGS",
                "doc_number": str(cell(row, "Номер") or "").strip() or None,
                "doc_total": _num(cell(row, "СуммаДокумента")),
                "warehouse": str(cell(row, "Склад") or "").strip() or None,
                "product": str(cell(row, "НоменклатураНаименование") or "").strip() or None,
                "qty": _num(cell(row, "Количество")),
            })
        else:
            amount = _num(cell(row, "Сумма"))
            if amount is None:
                continue
            out.append({
                "kind": kind, "date": d,
                "counterparty": str(cell(row, "Контрагент") or "").strip() or None,
                "amount": amount,
                "currency": str(cell(row, "Валюта") or "KGS").strip() or "KGS",
                "doc_number": str(cell(row, "Номер") or "").strip() or None,
                "operation": str(
                    cell(row, "ВидОперации") or cell(row, "Основание") or ""
                ).strip() or None,
            })
    return kind, out


def import_tax_workbook(db: Session, content: bytes, filename: str,
                        user_id: int, org: str = "hygiene") -> dict:
    """Импорт файла налоговой базы. Каждая загрузка заменяет данные своего
    вида целиком: выгрузки идут за всю историю, дедупликация не нужна.
    Используется и ручной кнопкой, и автоприёмом из папки Drive."""
    org = models.normalize_org(org)
    kind, parsed = _parse(content, filename or "")
    if not parsed:
        raise HTTPException(status_code=400, detail="В файле не нашлось ни одной строки с датой и суммой")
    # Снапшот-замена: сначала парсим (выше), только потом удаляем старое —
    # битый файл не может стереть данные.
    db.query(models.TaxOperation).filter(
        models.TaxOperation.organization == org,
        models.TaxOperation.kind == kind,
    ).delete(synchronize_session=False)
    for p in parsed:
        db.add(models.TaxOperation(organization=org, **p))
    db.add(models.ImportLog(
        filename=f"[налоговая:{org}:{kind}] {filename}",
        user_id=user_id, added=len(parsed), skipped=0, errors_count=0,
    ))
    db.commit()
    return {"kind": kind, "added": len(parsed)}


@router.post("/import")
async def tax_import(
    db: Session = Depends(get_db),
    user: models.User = Depends(can_edit),
    file: UploadFile = File(...),
    org: str = Form(default="hygiene"),
):
    content = await file.read()
    return import_tax_workbook(db, content, file.filename or "", user.id, org)


KIND_LABEL = {"sale": "Реализации", "return": "Возвраты",
              "cash_in": "Касса · приход", "cash_out": "Касса · расход"}


from pydantic import BaseModel


class TaxLinkItem(BaseModel):
    tax_name: str
    upr_names: list[str] | None = None  # пустой список/None — удалить связки


def _links_map(db: Session) -> dict[str, list[str]]:
    """Карта «налоговое имя → имена управленки» (их может быть несколько)."""
    out: dict[str, list[str]] = {}
    for l in db.query(models.TaxClientLink).all():
        out.setdefault(l.tax_name, []).append(l.upr_name)
    return out


@router.get("/links")
def tax_links(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """Контрагенты налоговой базы с оборотами + текущие связки с управленкой."""
    links = _links_map(db)
    agg: dict[str, dict] = {}
    for op in db.query(models.TaxOperation).all():
        if not op.counterparty:
            continue
        a = agg.setdefault(op.counterparty, {"amount": 0.0, "count": 0})
        a["amount"] += float(op.amount)
        a["count"] += 1
    # Кандидаты со стороны управленки: клиенты продаж, плательщики, контрагенты
    # расходов — всё, с чем налоговую операцию можно связать.
    upr: set[str] = set()
    for (c,) in db.query(models.Sale.client).distinct():
        upr.add(c)
    for (c,) in db.query(models.Receipt.payer).distinct():
        upr.add(c)
    for (c,) in db.query(models.Expense.counterparty).distinct():
        upr.add(c)
    return {
        "clients": sorted(
            ({"tax_name": name, "amount": round(v["amount"], 2),
              "count": v["count"], "upr_names": links.get(name, [])}
             for name, v in agg.items()),
            key=lambda x: -x["amount"],
        ),
        "upr_options": sorted(n for n in upr if n),
    }


@router.post("/links")
def tax_link_save(
    payload: TaxLinkItem,
    db: Session = Depends(get_db),
    _: models.User = Depends(can_edit),
):
    """Сохранить связки контрагента: список имён управленки заменяется целиком
    (пустой список — удалить все)."""
    tax_name = payload.tax_name.strip()
    if not tax_name:
        raise HTTPException(status_code=400, detail="Пустое имя контрагента")
    db.query(models.TaxClientLink).filter_by(tax_name=tax_name).delete(
        synchronize_session=False)
    names = [n.strip() for n in (payload.upr_names or []) if n and n.strip()]
    for n in dict.fromkeys(names):  # без дублей, порядок сохранён
        db.add(models.TaxClientLink(tax_name=tax_name, upr_name=n))
    db.commit()
    return {"status": "ok", "count": len(set(names))}


DOCS_CAP = 500


@router.get("/docs")
def tax_docs(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    kind: str = "sale",
    org: str = "all",
):
    """Реестр операций налогового контура — каждый документ строкой.

    Для реализаций строки файла собираются в документы (номер + дата +
    контрагент): итоговая сумма и число позиций. Остальные виды — одна строка
    файла и есть одна операция."""
    if kind not in KIND_LABEL:
        raise HTTPException(status_code=400, detail="Неизвестный вид операций")
    q = db.query(models.TaxOperation).filter(models.TaxOperation.kind == kind)
    o = models.normalize_org(org) if (org or "").lower() in models.ORGS else None
    if o:
        q = q.filter(models.TaxOperation.organization == o)
    rows = q.all()

    if kind == "sale":
        docs: dict = {}
        for r in rows:
            key = (r.doc_number, r.date, r.counterparty)
            d = docs.get(key)
            if d is None:
                d = docs[key] = {
                    "date": r.date.isoformat(),
                    "doc_number": r.doc_number,
                    "counterparty": r.counterparty,
                    "warehouse": r.warehouse,
                    "amount": 0.0,
                    "doc_total": None,
                    "positions": 0,
                    "currency": r.currency,
                }
            d["positions"] += 1
            d["amount"] += float(r.amount)
            if r.doc_total is not None:
                d["doc_total"] = float(r.doc_total)
        items = []
        for d in docs.values():
            amt = d.pop("doc_total") or d["amount"]
            items.append({**d, "amount": round(float(amt), 2)})
    else:
        items = [
            {
                "date": r.date.isoformat(),
                "doc_number": r.doc_number,
                "counterparty": r.counterparty,
                "operation": r.operation,
                "amount": round(float(r.amount), 2),
                "currency": r.currency,
            }
            for r in rows
        ]

    items.sort(key=lambda x: (x["date"], x.get("doc_number") or ""), reverse=True)

    # --- Пара в управленке для каждой операции ---
    # Контрагенты в контурах разные (в налоговой — юрлица, в управленке —
    # точки), поэтому имя ключом быть не может. Пара ищется по сумме (до
    # копеек) и близкой дате: сначала день в день, потом в пределах недели —
    # проводки в налоговую базу часто заносятся с отставанием.
    def upr_candidates() -> list[dict]:
        uq_org = o or "hygiene"  # налоговый контур пока только по Hygiene
        if kind == "sale":
            docs: dict = {}
            for s in db.query(models.Sale).filter(
                    models.Sale.organization == uq_org).all():
                key = (s.doc_number or f"~{s.client}", s.date, s.client)
                d = docs.setdefault(key, {"date": s.date, "who": s.client,
                                          "amount": 0.0, "doc_total": None,
                                          "currency": "KGS"})
                d["amount"] += float(s.amount) * (1 - float(s.discount_pct or 0) / 100)
                if s.doc_total is not None:
                    d["doc_total"] = float(s.doc_total)
            out = []
            for d in docs.values():
                out.append({"date": d["date"], "who": d["who"], "currency": "KGS",
                            "amount": round(float(d["doc_total"] or d["amount"]), 2)})
            return out
        if kind == "return":
            return [{"date": r.date, "who": r.client, "currency": r.currency,
                     "amount": round(float(r.amount), 2)}
                    for r in db.query(models.ReturnDoc).filter(
                        models.ReturnDoc.organization == uq_org).all()]
        if kind == "cash_in":
            return [{"date": r.date, "who": r.payer, "currency": r.currency,
                     "amount": round(float(r.amount), 2)}
                    for r in db.query(models.Receipt).filter(
                        models.Receipt.organization == uq_org).all()]
        return [{"date": e.date, "who": e.counterparty, "currency": e.currency,
                 "amount": round(float(e.amount), 2)}
                for e in db.query(models.Expense).filter(
                    models.Expense.organization == uq_org).all()]

    cands = upr_candidates()
    used: set = set()
    # Связки контрагентов НАЛ ↔ УПР: для связанных имя становится ключом
    # сверки — сначала ищем пару только среди операций связанных контрагентов
    # (окно шире, две недели), и лишь потом среди всех по сумме и дате.
    # Связанных имён может быть несколько (Императив → все точки Алдей).
    name_links = _links_map(db)

    def scan(item: dict, d0, cur, who: set | None, window: int):
        best, best_days = None, None
        for j, c in enumerate(cands):
            if j in used or (c["currency"] or "KGS") != cur:
                continue
            if who is not None and (c["who"] or "") not in who:
                continue
            if abs(c["amount"] - item["amount"]) >= 0.5:
                continue
            days = abs((c["date"] - d0).days)
            if days > window:
                continue
            if best is None or days < best_days:
                best, best_days = j, days
                if days == 0:
                    break
        return best, best_days

    def find_pair(item: dict):
        d0 = date.fromisoformat(item["date"])
        cur = item.get("currency") or "KGS"
        linked = set(name_links.get(item.get("counterparty") or "", []))
        best, best_days = (None, None)
        if linked:
            best, best_days = scan(item, d0, cur, linked, 14)
        if best is None:
            best, best_days = scan(item, d0, cur, None, 7)
        if best is None:
            return None
        used.add(best)
        c = cands[best]
        return {"date": c["date"].isoformat(), "who": c["who"], "days": best_days,
                "by_link": bool(linked and (c["who"] or "") in linked)}

    matched = 0
    for item in items:
        pair = find_pair(item)
        item["upr"] = pair
        if pair:
            matched += 1

    unmatched_amount = round(sum(i["amount"] for i in items if not i["upr"]), 2)
    return {
        "kind": kind,
        "label": KIND_LABEL[kind],
        "count": len(items),
        "amount": round(sum(i["amount"] for i in items), 2),
        "matched": matched,
        "unmatched": len(items) - matched,
        "unmatched_amount": unmatched_amount,
        "items": items[:DOCS_CAP],
        "cap": DOCS_CAP,
    }


@router.get("/compare")
def tax_compare(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    org: str = "all",
):
    """Сверка трёх контуров по месяцам: управленка ↔ налоговая ↔ SalesDoc.

    Поклиентно контуры не сопоставить — в налоговой базе продажи проведены на
    юрлица, а в управленке и SalesDoc клиенты — розничные точки. Поэтому
    сверяем агрегатами: выручка и поступления денег по месяцам. Это отвечает
    на два вопроса сразу: какая доля оборота проведена официально (НАЛ/УПР) и
    насколько SalesDoc совпадает с управленкой (Δ SD)."""
    from .receipts import CUSTOMER_PAYMENT_PREFIX
    from ..services import salesdoc as sd

    o = models.normalize_org(org) if (org or "").lower() in models.ORGS else None

    sales: dict[str, dict] = defaultdict(lambda: {"upr": 0.0, "nal": 0.0, "sd": 0.0})
    money: dict[str, dict] = defaultdict(lambda: {"upr": 0.0, "nal": 0.0, "sd": 0.0})

    def mk(d) -> str:
        return f"{d.year:04d}-{d.month:02d}"

    # --- Управленка: продажи документами (итог дока — один раз) ---
    q = db.query(models.Sale)
    if o:
        q = q.filter(models.Sale.organization == o)
    seen_docs: set = set()
    for s in q.all():
        if s.doc_number and s.doc_total is not None:
            key = (s.doc_number, s.date, s.client)
            if key in seen_docs:
                continue
            seen_docs.add(key)
            sales[mk(s.date)]["upr"] += float(s.doc_total)
        else:
            sales[mk(s.date)]["upr"] += float(s.amount) * (
                1 - float(s.discount_pct or 0) / 100)

    # --- Управленка: поступления от покупателей ---
    rq = db.query(models.Receipt)
    if o:
        rq = rq.filter(models.Receipt.organization == o)
    for r in rq.all():
        if r.operation.startswith(CUSTOMER_PAYMENT_PREFIX):
            money[mk(r.date)]["upr"] += float(r.amount_kgs)

    # --- Налоговая ---
    tq = db.query(models.TaxOperation)
    if o:
        tq = tq.filter(models.TaxOperation.organization == o)
    for t in tq.all():
        if t.kind == "sale":
            sales[mk(t.date)]["nal"] += float(t.amount)
        elif t.kind == "cash_in" and "покупател" in (t.operation or "").lower():
            money[mk(t.date)]["nal"] += float(t.amount)

    # --- SalesDoc (зеркало): отгрузки по складам фирмы + оплаты ---
    store_ids = None
    if o:
        rows = [s for s in db.query(models.SalesDocStore).all() if s.store_id]
        mine = {s.store_id.lower() for s in rows if s.organization == o}
        unmapped = {s.store_id.lower() for s in rows if not s.organization}
        store_ids = (mine | unmapped) if mine else None
    oq = db.query(models.SalesDocOrder).filter(
        models.SalesDocOrder.status.in_(sorted(sd.SHIPPED_STATUSES)))
    if store_ids:
        oq = oq.filter(models.SalesDocOrder.store_sd_id.in_(store_ids))
    for r in oq.all():
        if r.date:
            sales[mk(r.date)]["sd"] += float(r.amount or 0)
    for p in db.query(models.SalesDocPayment).filter(
            models.SalesDocPayment.txn == sd.PAYMENT_TXN).all():
        if p.date:
            money[mk(p.date)]["sd"] += float(p.amount or 0)

    def table(agg: dict) -> list[dict]:
        out = []
        for m in sorted(agg, reverse=True):
            v = agg[m]
            out.append({
                "month": m,
                "upr": round(v["upr"], 2),
                "nal": round(v["nal"], 2),
                "sd": round(v["sd"], 2),
                # Доля официально проведённого от управленческого оборота.
                "nal_share": round(v["nal"] / v["upr"] * 100, 1) if v["upr"] else None,
                "sd_diff": round(v["sd"] - v["upr"], 2),
            })
        return out

    def totals(agg: dict) -> dict:
        u = sum(v["upr"] for v in agg.values())
        n = sum(v["nal"] for v in agg.values())
        s_ = sum(v["sd"] for v in agg.values())
        return {"upr": round(u, 2), "nal": round(n, 2), "sd": round(s_, 2),
                "nal_share": round(n / u * 100, 1) if u else None,
                "sd_diff": round(s_ - u, 2)}

    return {
        "org": o or "all",
        # Оплаты SalesDoc по фирмам не делятся идеально (аванс без заказов
        # виден в обеих) — честно предупреждаем в интерфейсе.
        "sales": {"rows": table(sales), "totals": totals(sales)},
        "money": {"rows": table(money), "totals": totals(money)},
    }


@router.get("/summary")
def tax_summary(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    org: str = "all",
):
    """Сводка налогового контура: выручка по годам, касса по видам операций,
    подотчёт по людям. Всё считается из своей таблицы — управленка не трогается."""
    q = db.query(models.TaxOperation)
    o = models.normalize_org(org) if (org or "").lower() in models.ORGS else None
    if o:
        q = q.filter(models.TaxOperation.organization == o)
    ops = q.all()
    # Связки: обороты раздробленных юрлиц собираются под общим именем
    # управленки («Байго Трейд» вместо шести отдельных ИП). Склеиваем только
    # однозначные связки: если у налогового имени несколько управленческих
    # (Императив → точки Алдей), оборот остаётся под налоговым именем —
    # разделить его по точкам не по чему.
    name_links = _links_map(db)

    kinds: dict[str, dict] = {}
    sales_by_year: dict[int, float] = defaultdict(float)
    cash_by_op: dict[tuple, float] = defaultdict(float)
    podotchet: dict[str, dict] = {}
    clients: dict[str, float] = defaultdict(float)
    last_date: dict[str, str] = {}

    for op in ops:
        k = kinds.setdefault(op.kind, {"count": 0, "amount": 0.0})
        amt = float(op.amount)
        k["count"] += 1
        k["amount"] += amt
        if op.date and (op.kind not in last_date or op.date.isoformat() > last_date[op.kind]):
            last_date[op.kind] = op.date.isoformat()
        if op.kind == "sale":
            sales_by_year[op.date.year] += amt
            if op.counterparty:
                targets = name_links.get(op.counterparty) or []
                name = targets[0] if len(targets) == 1 else op.counterparty
                clients[name] += amt
        elif op.kind in ("cash_in", "cash_out"):
            cash_by_op[(op.kind, op.operation or "(без вида)")] += amt
            oper = (op.operation or "").lower()
            if "подотчет" in oper or "подотчёт" in oper:
                p = podotchet.setdefault(op.counterparty or "(не указан)",
                                         {"issued": 0.0, "returned": 0.0})
                if op.kind == "cash_out":
                    p["issued"] += amt
                else:
                    p["returned"] += amt

    return {
        "org": o or "all",
        "kinds": [
            {"kind": k, "label": KIND_LABEL.get(k, k),
             "count": v["count"], "amount": round(v["amount"], 2),
             "last_date": last_date.get(k)}
            for k, v in sorted(kinds.items())
        ],
        "sales_by_year": [
            {"year": y, "amount": round(a, 2)}
            for y, a in sorted(sales_by_year.items())
        ],
        "cash_by_operation": sorted(
            ({"direction": k[0], "operation": k[1], "amount": round(a, 2)}
             for k, a in cash_by_op.items()),
            key=lambda x: -x["amount"],
        ),
        "podotchet": sorted(
            ({"person": name, "issued": round(v["issued"], 2),
              "returned": round(v["returned"], 2),
              "hanging": round(v["issued"] - v["returned"], 2)}
             for name, v in podotchet.items()),
            key=lambda x: -x["hanging"],
        ),
        "top_clients": sorted(
            ({"client": c, "amount": round(a, 2)} for c, a in clients.items()),
            key=lambda x: -x["amount"],
        )[:15],
    }
