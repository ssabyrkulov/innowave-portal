"""Расход денежных средств: исходящие платёжки банка (ВыгрузкаППИсход) и
РКО кассы (ВыгрузкаРКО). Гибкий разбор: понимает и «плоский» формат
Дата/Сумма/Валюта/Контрагент, и расширенный с Документ/Номер/Основание.
"""

import hashlib
import io
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import get_current_user, require_roles
from .receipts import DEFAULT_RATES

router = APIRouter(prefix="/expenses", tags=["expenses"])

admin_only = require_roles(models.Role.admin)

HEADERS = {
    # дата
    "Дата": "date",
    "ДатаДокумента": "date",
    # сумма
    "Сумма": "amount",
    "СуммаДокумента": "amount",
    "СуммаПлатежа": "amount",
    "СуммаРасхода": "amount",
    # валюта
    "Валюта": "currency",
    "ВалютаДокумента": "currency",
    "ВалютаДокументаНаименование": "currency",
    # контрагент/получатель
    "Контрагент": "counterparty",
    "КонтрагентНаименование": "counterparty",
    "Получатель": "counterparty",
    "ПолучательНаименование": "counterparty",
    # основание/назначение/статья
    "Основание": "basis",
    "ВидОперации": "basis",
    "Назначение": "basis",
    "НазначениеПлатежа": "basis",
    "СтатьяДвиженияДенежныхСредств": "basis",
    "СтатьяДДС": "basis",
    "Статья": "basis",
    # номер
    "Номер": "doc_number",
    "НомерДокумента": "doc_number",
}


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.split()[0], "%d.%m.%Y").date()
        except ValueError:
            return None
    return None


def import_expenses_workbook(
    db: Session, content: bytes, filename: str, user_id: int, kind: str = "bank",
    org: str = models.DEFAULT_ORG,
) -> dict:
    org = models.normalize_org(org)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось открыть файл Excel")

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)

    header_idx, columns, buffered = None, {}, []
    seen_headers: list[str] = []
    for i, row in enumerate(rows_iter):
        buffered.append(row)
        if i >= 20:
            break
        cells = [str(c).strip() for c in row if c is not None]
        if cells and not seen_headers:
            seen_headers = cells
        matched = {
            j: HEADERS[str(c).strip()]
            for j, c in enumerate(row)
            if c is not None and str(c).strip() in HEADERS
        }
        # Минимум — дата и сумма; контрагент желателен, но не обязателен.
        if {"date", "amount"} <= set(matched.values()):
            header_idx, columns = i, matched
            seen_headers = cells
            break
    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Не найдены колонки расхода (нужны минимум Дата и Сумма). "
                   f"Заголовки в файле: {', '.join(seen_headers) or '—'}",
        )

    parsed, errors = [], []

    def process(row, line_no):
        data = {}
        for j, field in columns.items():
            # basis/doc_number могут встречаться в нескольких колонках — не
            # перетираем уже найденное непустое значение
            val = row[j] if j < len(row) else None
            if field in data and data[field]:
                continue
            data[field] = val
        if all(v is None for v in data.values()):
            return
        dt = _parse_date(data.get("date"))
        try:
            amount = float(str(data.get("amount")).replace(" ", "").replace(",", "."))
        except (TypeError, ValueError):
            amount = None
        cp = str(data.get("counterparty") or "").strip()
        if not dt or amount is None:
            if len(errors) < 20:
                errors.append(f"Строка {line_no}: нет даты или суммы")
            return
        currency = (str(data.get("currency") or "KGS").strip() or "KGS")[:3]
        rate = DEFAULT_RATES.get(currency, 1.0)
        parsed.append({
            "src_dt": str(data.get("date")),
            "date": dt,
            "amount": amount,
            "currency": currency,
            "rate": rate,
            "amount_kgs": round(amount * rate, 2),
            "counterparty": cp or "Не указан",
            "kind": kind,
            "basis": str(data.get("basis") or "").strip() or None,
            "doc_number": str(data.get("doc_number") or "").strip() or None,
        })

    line_no = header_idx + 1
    for line_no, row in enumerate(buffered[header_idx + 1:], start=header_idx + 2):
        process(row, line_no)
    for line_no, row in enumerate(rows_iter, start=line_no + 1):
        process(row, line_no)

    existing = {h for (h,) in db.query(models.Expense.row_hash).all()}
    seen: set[str] = set()
    occurrences: dict[str, int] = defaultdict(int)
    added = skipped = 0
    for p in parsed:
        base = "|".join(str(p[f]) for f in
                        ("src_dt", "amount", "currency", "counterparty", "kind", "doc_number"))
        occurrences[base] += 1
        h = hashlib.sha256(f"{org}|{base}#{occurrences[base]}".encode()).hexdigest()
        p = {k: v for k, v in p.items() if k != "src_dt"}
        if h in existing or h in seen:
            skipped += 1
            continue
        seen.add(h)
        db.add(models.Expense(**p, organization=org, row_hash=h))
        added += 1

    db.add(models.ImportLog(
        filename=f"[расход {kind}] {filename}",
        user_id=user_id,
        added=added,
        skipped=skipped,
        errors_count=len(errors),
    ))
    db.commit()
    return {
        "added": added,
        "skipped_duplicates": skipped,
        "errors": errors,
        "total_in_db": db.query(models.Expense).count(),
    }


@router.get("")
def list_expenses(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    kind: str | None = Query(default=None),
    limit: int = Query(default=300, le=1000),
    org: str = Query(default="all"),
):
    q = models.org_scope(db.query(models.Expense), models.Expense, org).order_by(models.Expense.date.desc())
    if kind:
        q = q.filter(models.Expense.kind == kind)
    return [
        {
            "id": e.id,
            "date": e.date.isoformat(),
            "amount": float(e.amount),
            "currency": e.currency,
            "amount_kgs": float(e.amount_kgs),
            "counterparty": e.counterparty,
            "kind": e.kind,
            "basis": e.basis,
        }
        for e in q.limit(limit).all()
    ]


@router.get("/summary")
def expenses_summary(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    org: str = Query(default="all"),
):
    q = models.org_scope(db.query(models.Expense), models.Expense, org)
    if date_from:
        q = q.filter(models.Expense.date >= date_from)
    if date_to:
        q = q.filter(models.Expense.date <= date_to)
    rows = q.all()
    by_cp: dict[str, float] = defaultdict(float)
    by_month: dict[str, float] = defaultdict(float)
    total = 0.0
    for e in rows:
        amt = float(e.amount_kgs)
        total += amt
        by_cp[e.counterparty] += amt
        by_month[e.date.strftime("%Y-%m")] += amt
    return {
        "total": round(total, 2),
        "count": len(rows),
        "monthly": [
            {"month": m, "amount": round(v, 2)} for m, v in sorted(by_month.items())
        ],
        "top_counterparties": sorted(
            ({"name": k, "amount": round(v, 2)} for k, v in by_cp.items()),
            key=lambda x: -x["amount"],
        )[:15],
    }


@router.delete("", status_code=204)
def clear_expenses(
    db: Session = Depends(get_db),
    _: models.User = Depends(admin_only),
):
    db.query(models.Expense).delete()
    db.commit()
