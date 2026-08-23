"""Открытие Excel, устойчивое к особенностям выгрузок 1С.

Единственная причина существования модуля: 1С ред. 1.7 кладёт в архив
`xl/SharedStrings.xml` с большой буквы, а openpyxl ищет строчную и падает с
KeyError — файл не открывается вовсе. Починка была написана для налогового
импортёра и жила только в нём; когда весь пакет выгрузок перешёл на новый
формат, тот же архив стали присылать и остальные виды документов.
"""
import io
import zipfile

import openpyxl


def load_workbook(content: bytes, *, read_only: bool = True):
    """openpyxl.load_workbook с переупаковкой архива, если он нестандартный."""
    try:
        return openpyxl.load_workbook(
            io.BytesIO(content), data_only=True, read_only=read_only)
    except KeyError:
        return openpyxl.load_workbook(
            io.BytesIO(repack(content)), data_only=True, read_only=read_only)


def repack(content: bytes) -> bytes:
    """Пересобрать xlsx, приведя имя файла общих строк к ожидаемому."""
    src = zipfile.ZipFile(io.BytesIO(content))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            name = item.filename
            if name.lower().endswith("sharedstrings.xml"):
                name = "xl/sharedStrings.xml"
            dst.writestr(name, src.read(item.filename))
    return buf.getvalue()
